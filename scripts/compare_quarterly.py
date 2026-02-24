"""
Quarter-by-quarter comparison between pipeline output and human-made accounting records.
Produces a side-by-side difference table for 2020, 2021, 2022, 2024.

Usage: python3 scripts/compare_quarterly.py [year]
       year: optional, one of 2020 2021 2022 2024 (default: all)
"""

import os
import sys
import openpyxl
from collections import defaultdict

# ─── Paths ───────────────────────────────────────────────────────────────────
# BASE is the veritas-accounting project root (parent of scripts/).
# Works correctly regardless of which directory the script is called from.
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Human account books live in BOOKS/ inside the project root.
# Expected filenames (must match exactly):
#   BOOKS/2020Veritas China Account book.xlsx
#   BOOKS/2021Veritas China Account book.xlsx
#   BOOKS/2022Veritas China Account book.xlsx
#   BOOKS/2024Veritas China Account book.xlsx
# See docs/如何运行-how-to-run.md §五 for placement instructions.
_BOOKS_DIR = os.path.join(BASE, "BOOKS")
BOOKS = {
    2020: os.path.join(_BOOKS_DIR, "2020Veritas China Account book.xlsx"),
    2021: os.path.join(_BOOKS_DIR, "2021Veritas China Account book.xlsx"),
    2022: os.path.join(_BOOKS_DIR, "2022Veritas China Account book.xlsx"),
    2024: os.path.join(_BOOKS_DIR, "2024Veritas China Account book.xlsx"),
}

# ─── Pipeline ledger account names (as they appear in quarterly_report.xlsx) ─
PIPELINE_ACCOUNTS = [
    ("捐赠收入",   "4020"),
    ("投资收益",   "4030"),
    ("收入SC",    "4200"),
    ("收入OL",    "4300"),
    ("管理费用",   "5000"),
    ("组委工资",   "5010"),
    ("宣传推广",   "5020"),
    ("社群维护",   "5030"),
    ("支出SC运营", "5203"),
    ("支出OL讲师", "5301"),
    ("银行存款",   "1100"),
    ("应付税款",   "2209"),
    ("短期借款",   "2010"),
]

# ─── Human report row labels → pipeline account name ─────────────────────────
# Each entry: (search_label, pipeline_account, column_offset, negate)
# column_offset: 0=same col, 1=next col (for sub-items with label in prev col)
# negate: True if human sign is opposite to what we want (expenses shown positive)
HUMAN_MAP = [
    ("捐赠收入",     "捐赠收入",   0, False),
    ("投资收入",     "投资收益",   0, False),
    ("投资收益",     "投资收益",   0, False),   # 2024 style
    ("组委工资",     "组委工资",   0, True),
    ("组织管理费用", "管理费用",   0, True),
    ("    组织管理费用", "管理费用", 0, True),
]

# ─── Load pipeline quarterly totals ──────────────────────────────────────────
def load_pipeline(year):
    """Returns {(account_name, quarter): (cr, dr, net)} """
    # Prefer ledger_output.xlsx "Quarterly Report" sheet (always freshly generated)
    path = f"{BASE}/output/{year}/ledger_output.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Quarterly Report"]
    results = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        row = list(row) + [None] * 9
        lid, code, name, quarter, yr, cr, dr, net, cnt = row[:9]
        if name and quarter and isinstance(quarter, int) and isinstance(name, str):
            key = (name, quarter)
            results[key] = (cr or 0, dr or 0, net or 0)
    return results

# ─── Parse human report sheets ───────────────────────────────────────────────
def _col_val(row, col_idx):
    """Safely get value from row at col_idx."""
    try:
        v = row[col_idx]
        if isinstance(v, str) and v.startswith('#'):
            return None
        return v
    except IndexError:
        return None

def _find_quarter_cols(ws):
    """
    Find the column indices for each quarter label in the header row.
    Returns {quarter_label: col_index_0based}.
    E.g. {'2020Q1': 2, '2020Q2': 4, ...}
    """
    result = {}
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
        for ci, val in enumerate(row):
            if isinstance(val, str) and 'Q' in val and any(
                    str(y) in val for y in range(2019, 2025)):
                # e.g. '2020Q1', '2021Q2-3', '2022Q4'
                label = val.strip()
                if label not in result:
                    result[label] = ci
        if result:
            break
    return result

def load_human_2020(wb):
    """
    Returns {quarter: {account: net_value}} for Q1-Q4.
    2020 Q4 sheet has all quarters in one wide sheet.
    """
    out = defaultdict(dict)
    sheet_map = {
        "报表小结4Q": {
            "2020Q1": 2, "2020Q2": 4, "2020Q3": 8, "2020Q4": 10
        }
    }
    ws = wb["报表小结4Q"]
    for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
        label = row[0] if row else None
        if not label or not isinstance(label, str):
            continue
        label = label.strip()
        # Direct account rows
        for hlabel, paccount, offset, negate in HUMAN_MAP:
            if label == hlabel:
                for qname, ci in {"2020Q1": 2, "2020Q2": 4, "2020Q3": 8, "2020Q4": 10}.items():
                    v = _col_val(row, ci + offset)
                    if v is not None and isinstance(v, (int, float)):
                        qnum = int(qname[-1])
                        val = -v if negate else v
                        out[qnum][paccount] = out[qnum].get(paccount, 0) + val
        # 线上课程 学费收入 → 收入OL (net, includes 讲师工资)
        if label == "    线上课程":
            for qname, ci in {"2020Q1": 2, "2020Q2": 4, "2020Q3": 8, "2020Q4": 10}.items():
                v = _col_val(row, ci)
                if v is not None and isinstance(v, (int, float)):
                    qnum = int(qname[-1])
                    out[qnum]["收入OL_net"] = v
        # 唯理工作坊 → 收入SC net
        if "工作坊" in label or "唯理书院" in label or "书院" in label:
            if label.startswith("    "):
                for qname, ci in {"2020Q1": 2, "2020Q2": 4, "2020Q3": 8, "2020Q4": 10}.items():
                    v = _col_val(row, ci)
                    if v is not None and isinstance(v, (int, float)):
                        qnum = int(qname[-1])
                        out[qnum]["收入SC_net"] = out[qnum].get("收入SC_net", 0) + v
    return out

def load_human_2021(wb):
    """2021: Q1 from 报表小结1Q col5, Q2-3 incremental from 2-3Q col7, Q4 from 4Q."""
    out = defaultdict(dict)

    def _extract_from_sheet(sheet_name, col_map):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
            label = row[0] if row else None
            if not label or not isinstance(label, str):
                continue
            label = label.strip()
            for hlabel, paccount, offset, negate in HUMAN_MAP:
                if label == hlabel:
                    for qname, ci in col_map.items():
                        v = _col_val(row, ci + offset)
                        if v is not None and isinstance(v, (int, float)):
                            val = -v if negate else v
                            out[qname][paccount] = val
            if label == "    线上课程" or label == "线上课程":
                for qname, ci in col_map.items():
                    v = _col_val(row, ci)
                    if v is not None and isinstance(v, (int, float)):
                        out[qname]["收入OL_net"] = v
            if "工作坊" in label and label.startswith("    "):
                for qname, ci in col_map.items():
                    v = _col_val(row, ci)
                    if v is not None and isinstance(v, (int, float)):
                        out[qname]["收入SC_net"] = out[qname].get("收入SC_net", 0) + v

    _extract_from_sheet("报表小结1Q", {"Q1": 4})
    _extract_from_sheet("报表小结2-3Q", {"Q2-3": 6})
    _extract_from_sheet("报表小结4Q", {"Q4": 10})  # if valid; 2021 Q4 has #REF! issues
    return out

def load_human_2022(wb):
    """2022: Q1 from 1Q sheet col3, Q2 col5, Q3 col7, Q4 incremental."""
    out = defaultdict(dict)

    def _extract_from_sheet(sheet_name, col_map):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
            label = row[0] if row else None
            if not label or not isinstance(label, str):
                continue
            label = label.strip()
            for hlabel, paccount, offset, negate in HUMAN_MAP:
                if label == hlabel:
                    for qname, ci in col_map.items():
                        v = _col_val(row, ci + offset)
                        if v is not None and isinstance(v, (int, float)):
                            val = -v if negate else v
                            out[qname][paccount] = val
            if label == "    线上课程" or label == "线上课程":
                for qname, ci in col_map.items():
                    v = _col_val(row, ci)
                    if v is not None and isinstance(v, (int, float)):
                        out[qname]["收入OL_net"] = v
            if ("工作坊" in label or "书院" in label) and label.startswith("    "):
                for qname, ci in col_map.items():
                    v = _col_val(row, ci)
                    if v is not None and isinstance(v, (int, float)):
                        out[qname]["收入SC_net"] = out[qname].get("收入SC_net", 0) + v

    # 2022 sheets use odd-numbered columns (merged-cell layout): 1=2021FY, 3=2022Q1,
    # 5=2022Q2, 7=2022Q3, 9=2022Q2-3, 11=2022Q4.
    _extract_from_sheet("报表小结1Q",   {"Q1": 3})
    _extract_from_sheet("报表小结2-3Q", {"Q2": 5, "Q3": 7})
    _extract_from_sheet("报表小结4Q",   {"Q4": 11})
    return out

def load_human_2024(wb):
    """2024: Q4 from 报表Q4 sheet (has per-account CR/DR), FY from 报表小结FY."""
    out = defaultdict(dict)

    # 报表Q4: account-level CR/DR
    if "报表Q4" in wb.sheetnames:
        ws = wb["报表Q4"]
        # Row format: (name, code, cr, dr, net, ...)
        acct_map = {
            "捐赠收入": "捐赠收入",
            "投资收益": "投资收益",
            "收入SC": "收入SC",
            "收入OL": "收入OL",
            "管理费用": "管理费用",
            "组委工资": "组委工资",
            "宣传推广": "宣传推广",
            "社群维护": "社群维护",
            "支出SC运营": "支出SC运营",
            "支出OL讲师": "支出OL讲师",
            "银行存款": "银行存款",
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row else None
            if not name or not isinstance(name, str):
                continue
            paccount = acct_map.get(name.strip())
            if paccount:
                cr = row[2] if len(row) > 2 else None
                dr = row[3] if len(row) > 3 else None
                net = row[4] if len(row) > 4 else None
                out["Q4"][paccount] = {
                    "cr": cr or 0,
                    "dr": dr or 0,
                    "net": net or 0,
                }

    return out

# ─── Format helpers ───────────────────────────────────────────────────────────
def fmt(v):
    if v is None:
        return "       —"
    return f"{v:>10,.2f}"

def fmt_diff(d):
    if d is None:
        return "       —"
    if abs(d) < 0.01:
        return "      ✓ "
    return f"{d:>+10,.2f}"

# ─── Compare pipeline vs human ────────────────────────────────────────────────
def compare_year(year):
    print(f"\n{'='*70}")
    print(f"  YEAR {year}")
    print(f"{'='*70}")

    try:
        pipeline = load_pipeline(year)
    except FileNotFoundError:
        print(f"  No pipeline output for {year}")
        return

    wb = openpyxl.load_workbook(BOOKS[year], data_only=True)

    if year == 2020:
        human = load_human_2020(wb)
        quarters = [1, 2, 3, 4]
    elif year == 2021:
        human = load_human_2021(wb)
        quarters = ["Q1", "Q2-3", "Q4"]
    elif year == 2022:
        human = load_human_2022(wb)
        quarters = ["Q1", "Q2", "Q3", "Q4"]
    elif year == 2024:
        human = load_human_2024(wb)
        quarters = ["Q4"]

    for q in quarters:
        # Map quarter label to pipeline quarter number
        if isinstance(q, int):
            pq = q
        elif q == "Q1":
            pq = 1
        elif q in ("Q2-3",):
            pq = None  # combined, skip for now
        elif q == "Q2":
            pq = 2
        elif q == "Q3":
            pq = 3
        elif q == "Q4":
            pq = 4

        print(f"\n  --- {year} {q} ---")
        hdata = human.get(q, {})

        # Build comparison table for directly-comparable accounts
        header = f"  {'Account':<18} {'Human':>12} {'Pipeline CR':>12} {'Pipeline DR':>12} {'Pipeline Net':>12} {'Diff(Net)':>10}"
        print(header)
        print("  " + "-"*76)

        for acct, code in PIPELINE_ACCOUNTS:
            h_val = hdata.get(acct)

            if pq is not None:
                p_data = pipeline.get((acct, pq), (None, None, None))
                p_cr, p_dr, p_net = p_data
            else:
                p_cr = p_dr = p_net = None

            # For 2024 Q4, human data is a dict with cr/dr/net
            if isinstance(h_val, dict):
                h_net = h_val.get("net")
                h_cr  = h_val.get("cr")
                h_dr  = h_val.get("dr")
            else:
                h_net = h_val
                h_cr  = h_dr = None

            # Skip if all None
            if h_net is None and p_net is None:
                continue

            diff = None
            if h_net is not None and p_net is not None:
                diff = p_net - h_net

            # For expense accounts, human shows them as positive (already negated above)
            # Pipeline net = CR - DR; for expense accounts net is usually negative
            print(f"  {acct:<18} {fmt(h_net)} {fmt(p_cr)} {fmt(p_dr)} {fmt(p_net)} {fmt_diff(diff)}")

        # Also show 收入OL_net and 收入SC_net from human (project net)
        for key in ("收入OL_net", "收入SC_net"):
            h_net = hdata.get(key)
            if h_net is not None:
                label = key.replace("_net", " (proj.net)")
                if pq:
                    ol = pipeline.get(("收入OL", pq), (0,0,0))
                    sc = pipeline.get(("收入SC", pq), (0,0,0))
                    if key == "收入OL_net":
                        p_net = ol[2]
                    else:
                        p_net = sc[2]
                    diff = p_net - h_net if p_net is not None else None
                else:
                    p_net = None
                    diff = None
                print(f"  {label:<18} {fmt(h_net)} {'':>12} {'':>12} {fmt(p_net)} {fmt_diff(diff)}")

        # Show any pipeline accounts not in human data
        if pq:
            print(f"\n  Pipeline-only accounts (no human equivalent found):")
            all_pipeline_names = {a for a, _ in PIPELINE_ACCOUNTS}
            for (name, quarter), (cr, dr, net) in pipeline.items():
                if quarter == pq and name not in all_pipeline_names and abs(net or 0) > 0.01:
                    print(f"    {name:<20} CR={fmt(cr)} DR={fmt(dr)} Net={fmt(net)}")

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    years_arg = sys.argv[1:] if len(sys.argv) > 1 else None
    if years_arg:
        years = [int(y) for y in years_arg]
    else:
        years = [2020, 2021, 2022, 2024]

    for year in years:
        compare_year(year)
