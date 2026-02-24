"""
Update 账目分类明细.xlsx:
1. Resolve duplicate (year, type) rule pairs by keeping the correct single row
2. Add missing rules for types used in Journal-RMB but currently absent

Run: python3 scripts/update_rules.py
Writes to 账目分类明细_updated.xlsx (review before replacing original).
"""

import shutil
import os
import openpyxl

RULES_FILE = "/Users/JerryChen/proj/cursor/veritas-accounting/账目分类明细.xlsx"
OUT_FILE   = "/Users/JerryChen/proj/cursor/veritas-accounting/账目分类明细_updated.xlsx"

# ─── Rows to DELETE (1-based row numbers in journal_to_ledger sheet) ──────────
# Decision rationale documented beside each entry.
ROWS_TO_DELETE = set([
    # (2020, 'OL') ×3 → keep row 149 (cr=提供服务收入4301 dr=银行存款 = correct income direction)
    76,   # cr=银行存款 dr=提供服务收入 = wrong direction (T- income = refund direction)
    148,  # cr=其他应付款 dr=银行存款 = deposit type, wrong for 'OL' which is pure income in 2020

    # (2020, 'WX-OL') ×2 → keep row 155 (cr=提供服务收入4301 dr=银行存款 = correct income)
    154,  # cr=其他应付款 dr=银行存款 = deposit type; WX-OL is WeChat-collected income, not deposit

    # (2020, 'Workshop-WAGE') ×2 → keep row 29 (cr=银行存款 dr=业务活动成本5101 = expense correct)
    174,  # cr=None dr=None = empty rule, useless

    # (2021, 'Workshop') ×2 → keep row 152 (cr=工作坊收入4201 dr=银行存款 = correct income)
    63,   # cr=银行存款 dr=工作坊收入 = wrong direction (T- direction = opposite)

    # (2021, 'workshop-FEE') ×2 → keep row 71 (cr=银行存款 dr=工作坊运营成本5101 = specific)
    17,   # cr=银行存款 dr=业务活动成本5001 = old generic ledger code, superseded by 5101

    # (2022, 'OL') ×8 → keep row 108 (cr=OL收入4301 dr=银行存款 = correct income direction)
    4,    # cr=銀行存款 dr=OL收入 = wrong direction
    5,    # same wrong direction, duplicate of row 4
    12,   # cr=銀行存款 dr=OL讲师工资5401 = lecturer pay — should use '工资' type, not 'OL'
    41,   # cr=銀行存款 dr=其他應付款2209 = deposit — should use 'OL-D' type, not 'OL'
    109,  # exact duplicate of row 108 (cr=OL收入4301 dr=銀行存款)
    114,  # cr=其他應付款 dr=銀行存款 = deposit type, should be 'OL-D'
    115,  # same as 114, duplicate

    # (2022, 'OL-N') ×2 → keep row 121 (cr=OL收入4301 dr=银行存款 = correct income)
    8,    # cr=銀行存款 dr=OL收入 = wrong direction

    # (2022, '书院') ×5 → keep row 125 (cr=工作坊/书院收入4201 dr=銀行存款 = correct income)
    53,   # cr=銀行存款 dr=工作坊/书院收入 = wrong direction (T-)
    58,   # cr=銀行存款 dr=工作坊/书院讲师工资 = lecturer pay — 书院 type is income, not wages
    61,   # cr=銀行存款 dr=工作坊/书院运营成本 = ops expense — wrong for income type
    124,  # cr=OL收入 dr=銀行存款 = wrong account (OL vs SC) and wrong direction; note: "还未建立这个分类"

    # (2022, '工资') ×3 → keep row 100 (cr=銀行存款 dr=组委工资5601 = correct for staff wages)
    59,   # cr=銀行存款 dr=工作坊/书院讲师工资5301 new_type=支出SC组委 — less accurate account
    60,   # cr=銀行存款 dr=工作坊/书院讲师工资5301 new_type=支出SC讲师 — duplicate of 59 for different sub-type

    # (2022, '报销') ×3 → keep row 82 (cr=銀行存款 dr=管理費用5201 = correct for reimbursements)
    18,   # cr=銀行存款 dr=业务活动成本5001 = old generic code, superseded
    62,   # cr=銀行存款 dr=工作坊/书院运营成本5101 = SC operations, wrong for general reimbursements

    # (2022, '赞赏') ×2 → keep row 143 (cr=捐贈收入4101 dr=銀行存款 = WeChat tips = donations)
    142,  # cr=投資收益 dr=銀行存款 = investment income, wrong for appreciation tips

    # (2022, '运营') ×2 → keep row 86 (cr=銀行存款 dr=管理費用5201 = current code)
    21,   # cr=銀行存款 dr=业务活动成本5001 = old code, superseded

    # (2023, '/') ×2 → keep row 107 (internal transfer marker)
    171,  # duplicate with reversed direction; internal transfers cancel out

    # (2023, 'OL') ×7 → keep row 110 (cr=OL收入4301 dr=銀行存款 = correct income)
    6,    # cr=銀行存款 dr=OL收入 = wrong direction
    42,   # cr=銀行存款 dr=其他應付款 = deposit type, should use 'OL-D'
    111,  # exact duplicate of row 110
    116,  # cr=其他應付款 dr=銀行存款 = deposit type, should use 'OL-D'
    117,  # same as 116, duplicate
    120,  # cr=工作坊/书院收入 dr=銀行存款 = SC income, wrong for OL type

    # (2023, 'OL-D&OL助学金') ×2 → keep row 47 (deposit → 其他應付款)
    48,   # 奖助 (scholarship grant) — same old account code, deposit rule covers both in old system

    # (2023, 'OL-N') ×5 → keep row 122 (cr=OL收入4301 dr=銀行存款 = correct income)
    9,    # cr=銀行存款 dr=OL收入 = wrong direction
    49,   # cr=銀行存款 dr=其他應付款 = deposit, wrong for OL-N (recording income)
    50,   # duplicate of 49
    52,   # cr=銀行存款 dr=工作坊/书院收入 = SC income, wrong for OL type

    # (2023, '书院') ×4 → keep row 126 (cr=工作坊/书院收入4201 dr=銀行存款 = correct income)
    54,   # cr=銀行存款 dr=工作坊/书院收入 = wrong direction (T-)
    65,   # cr=銀行存款 dr=工作坊讲师工资 = lecturer pay, wrong for income type
    72,   # cr=銀行存款 dr=工作坊运营成本 = ops expense, wrong for income type

    # (2023, '书院-组委讲师') ×2 → keep row 66 (cr=銀行存款 dr=工作坊讲师工资5301 = wages)
    74,   # cr=銀行存款 dr=工作坊运营成本5101 = ops cost, less specific than wages account

    # (2023, '书院-酒店') ×2 → keep row 69 (cr=銀行存款 dr=工作坊讲师工资5301 = accommodation = compensation)
    128,  # T-支出SC运营 = reverse direction; accommodation should be normal expense direction

    # (2023, '运营') ×2 → keep row 87 (cr=銀行存款 dr=管理費用5201 = current code)
    22,   # cr=銀行存款 dr=业务活动成本5001 = old code

    # (2024, 'OL') ×8 → keep row 112 (cr=OL收入4301 dr=銀行存款 = correct income, Q1-Q3 use 'OL' for tuition)
    7,    # cr=銀行存款 dr=OL收入 = wrong direction
    13,   # cr=銀行存款 dr=OL讲师工资 = lecturer pay, should use separate type
    43,   # cr=銀行存款 dr=其他應付款 = deposit, should use 'OL-D' type
    98,   # cr=銀行存款 dr=组委工资 = org wages, wrong for OL income type
    113,  # exact duplicate of row 112
    118,  # cr=其他應付款 dr=銀行存款 = deposit, should use 'OL-D'
    119,  # duplicate of 118

    # (2024, '书院') ×3 → keep row 127 (cr=工作坊/书院收入4201 dr=銀行存款 = correct income)
    35,   # cr=銀行存款 dr=书院运营成本 = ops expense, wrong for income type
    55,   # cr=銀行存款 dr=工作坊/书院收入 = wrong direction (T-)

    # (2024, '书院住宿') ×3 → keep row 32 (cr=銀行存款 dr=书院讲师组委工资5301 = accommodation = compensation)
    36,   # cr=銀行存款 dr=书院运营成本5101 = ops, less accurate for accommodation
    129,  # T- reverse direction

    # (2024, '书院场地') ×2 → keep row 37 (cr=銀行存款 dr=书院运营成本5101 = venue expense)
    130,  # T- reverse direction

    # (2024, '书院学费') ×4 → keep row 132 (cr=工作坊/书院收入4201 dr=銀行存款 = correct income)
    57,   # cr=銀行存款 dr=工作坊/书院收入 = wrong direction (T-)
    99,   # cr=銀行存款 dr=组委工资 = completely wrong; note: "可能原分类是错误分类"
    131,  # cr=书院讲师组委工资 dr=銀行存款 = also flagged as possible error

    # (2024, '换汇-3000刀') ×2 → keep row 51 (cr=銀行存款 dr=其他貨幣資金1009 = correct for currency exchange)
    170,  # cr=銀行存款 dr=非限定性淨資產 = equity change, wrong for currency exchange

    # (2024, '运营') ×3 → keep row 88 (cr=銀行存款 dr=管理費用5201 = current code)
    23,   # cr=銀行存款 dr=业务活动成本5001 = old code
    145,  # T-管理费用 = reverse direction; regular ops expense goes in normal direction
])

# ─── New rows to ADD (appended to journal_to_ledger sheet) ───────────────────
# Format: (year, old_type, cr_ledger, dr_ledger, new_type, new_cr, new_dr, description)
NEW_RULES = [
    # ── 2024 missing types found in Journal-RMB Q4 ──────────────────────────
    (2024, "捐赠收入",     "捐贈收入4020",          "銀行存款1002",          "捐赠收入",     "捐赠收入",    "银行存款",     "捐赠/打赏收入"),
    (2024, "社群维护",     "銀行存款1002",          "社群维护5030",          "社群维护",     "银行存款",    "社群维护",     "社群活动/维护费用"),
    (2024, "宣传推广",     "銀行存款1002",          "宣传推广5020",          "宣传推广",     "银行存款",    "宣传推广",     "宣传/推广费用"),
    (2024, "应付OL押金",   "应付OL押金2301",        "銀行存款1002",          "应付OL押金",   "应付OL押金",  "银行存款",     "收取线上课程押金"),
    (2024, "应付OL奖助",   "应付OL奖助2302",        "銀行存款1002",          "应付OL奖助",   "应付OL奖助",  "银行存款",     "发放线上课程奖助学金"),
    (2024, "支出OL讲师",   "銀行存款1002",          "支出OL讲师5301",        "支出OL讲师",   "银行存款",    "支出OL讲师",   "支付线上课程讲师工资"),
    (2024, "T-收入OL",     "銀行存款1002",          "OL收入4301",            "T-收入OL",    "银行存款",    "收入OL",      "线上课程退款（收入冲回）"),
    (2024, "律师费",       "銀行存款1002",          "管理費用5000",          "律师费",       "银行存款",    "管理费用",     "律师费/法律服务费"),
    # ── 2024 missing types found in Journal-RMB Q1-Q3 ───────────────────────
    (2024, "OL-D",         "应付OL押金2301",        "銀行存款1002",          "应付OL押金",   "应付OL押金",  "银行存款",     "收取/退还OL押金"),
    (2024, "OL-奖学金",    "应付OL奖助2302",        "銀行存款1002",          "应付OL奖助",   "应付OL奖助",  "银行存款",     "发放/退还OL奖助学金"),
    (2024, "书院报销",     "銀行存款1002",          "书院运营成本5101",      "支出SC运营",   "银行存款",    "支出SC运营",   "书院运营报销"),
    (2024, "书院运营",     "銀行存款1002",          "书院运营成本5101",      "支出SC运营",   "银行存款",    "支出SC运营",   "书院运营费用"),
    (2024, "书院工资",     "銀行存款1002",          "书院讲师组委工资5301",  "支出SC讲师",   "银行存款",    "支出SC讲师",   "书院讲师/组委工资"),
    (2024, "OL-讲师工资",  "銀行存款1002",          "支出OL讲师5301",        "支出OL讲师",   "银行存款",    "支出OL讲师",   "OL讲师工资"),
    (2024, "OL押金",       "应付OL押金2301",        "銀行存款1002",          "应付OL押金",   "应付OL押金",  "银行存款",     "OL押金收取/退还"),
    (2024, "工资",         "銀行存款1002",          "组委工资5601",          "组委工资",     "银行存款",    "组委工资",     "工作人员工资（OL/SC组委）"),
    (2024, "研讨屋",       "銀行存款1002",          "播客/研讨屋5501",       "支出内容运营", "银行存款",    "支出内容运营", "研讨屋活动费用"),
    (2024, "播客",         "銀行存款1002",          "播客/研讨屋5501",       "支出内容运营", "银行存款",    "支出内容运营", "播客制作费用"),
    # ── 余利宝收益发放 for all years (currently only 2022 has it) ────────────
    (2020, "余利宝收益发放", "投資收益4030",         "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "余利宝每日收益"),
    (2021, "余利宝收益发放", "投資收益4030",         "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "余利宝每日收益"),
    (2023, "余利宝收益发放", "投資收益4030",         "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "余利宝每日收益"),
    (2024, "余利宝收益发放", "投資收益4030",         "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "余利宝每日收益"),
    # ── 基金收益 for 2020/2021 (older name for 余利宝收益发放) ─────────────
    (2020, "基金收益",     "投資收益4030",          "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "货币基金每日收益"),
    (2021, "基金收益",     "投資收益4030",          "銀行存款1002",          "投资收益",    "投资收益",    "银行存款",     "货币基金每日收益"),
]

def main():
    print(f"Loading {RULES_FILE} ...")
    shutil.copy2(RULES_FILE, OUT_FILE)
    wb = openpyxl.load_workbook(OUT_FILE)
    ws = wb["journal_to_ledger"]

    total_rows = ws.max_row
    print(f"  journal_to_ledger has {total_rows} rows")

    # ── Step 1: Delete rows (in reverse order to preserve row numbers) ────────
    rows_to_delete_sorted = sorted(ROWS_TO_DELETE, reverse=True)
    deleted = 0
    for row_num in rows_to_delete_sorted:
        # Verify the row still has content (sanity check)
        year_val = ws.cell(row=row_num, column=1).value
        type_val = ws.cell(row=row_num, column=2).value
        cr_val   = ws.cell(row=row_num, column=3).value
        dr_val   = ws.cell(row=row_num, column=4).value
        ws.delete_rows(row_num, 1)
        deleted += 1
        print(f"  Deleted row {row_num}: year={year_val} type={type_val!r} cr={cr_val} dr={dr_val}")

    print(f"\n  Deleted {deleted} rows. Remaining: {ws.max_row}")

    # ── Step 2: Append new rules ──────────────────────────────────────────────
    added = 0
    for rule in NEW_RULES:
        year, old_type, cr_ledger, dr_ledger, new_type, new_cr, new_dr, desc = rule

        # Check if rule already exists for this (year, old_type)
        exists = False
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == float(year) and row[1] == old_type:
                exists = True
                break
        if exists:
            print(f"  SKIP (exists): {year} {old_type!r}")
            continue

        # Append row
        new_row = [year, old_type, cr_ledger, dr_ledger, new_type, new_cr, new_dr, desc,
                   None, None, None, None, None, None, None, None, None, None,
                   None, None, None, None, None, None, None, None]
        ws.append(new_row)
        added += 1
        print(f"  Added: {year} {old_type!r} → cr={cr_ledger} dr={dr_ledger}")

    print(f"\n  Added {added} new rules. Final row count: {ws.max_row}")

    # ── Step 3: Verify no duplicates remain ───────────────────────────────────
    from collections import Counter
    combos = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            combos.append((row[0], row[1]))
    dups = [(k, v) for k, v in Counter(combos).items() if v > 1]
    if dups:
        print(f"\n  WARNING: {len(dups)} duplicate (year, type) pairs remain:")
        for k, v in dups:
            print(f"    {k}: {v} occurrences")
    else:
        print(f"\n  ✓ No duplicate (year, type) pairs remain!")

    wb.save(OUT_FILE)
    print(f"\nSaved to {OUT_FILE}")
    print("Review the file, then run:")
    print(f"  mv '{OUT_FILE}' '{RULES_FILE}'")

if __name__ == "__main__":
    main()
