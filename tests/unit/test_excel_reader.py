"""Unit tests for ExcelReader class."""

import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from veritas_accounting.excel.reader import ExcelReader
from veritas_accounting.utils.exceptions import ExcelIOError


class TestExcelReader:
    """Test suite for ExcelReader class."""

    def test_read_file_basic(self):
        """Test reading a basic Excel file."""
        reader = ExcelReader()

        # Create a temporary Excel file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with test data
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Amount"
            ws["A2"] = "Test Entry"
            ws["B2"] = 100.50
            wb.save(tmp_path)
            wb.close()

            # Read the file
            df = reader.read_file(tmp_path)

            # Verify data
            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            assert df.columns.tolist() == ["Name", "Amount"]
            assert df.iloc[0]["Name"] == "Test Entry"
            assert df.iloc[0]["Amount"] == 100.50

        finally:
            # Clean up
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_file_with_chinese_text(self):
        """Test reading Excel file with Chinese text (UTF-8 encoding)."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with Chinese text
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "描述"
            ws["B1"] = "金额"
            ws["A2"] = "测试条目"
            ws["B2"] = 200.75
            wb.save(tmp_path)
            wb.close()

            # Read the file
            df = reader.read_file(tmp_path)

            # Verify Chinese text is read correctly
            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            assert df.columns.tolist() == ["描述", "金额"]
            assert df.iloc[0]["描述"] == "测试条目"
            assert df.iloc[0]["金额"] == 200.75

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_file_missing_file(self):
        """Test reading a non-existent file raises ExcelIOError."""
        reader = ExcelReader()
        non_existent = Path("/nonexistent/file.xlsx")

        with pytest.raises(ExcelIOError) as exc_info:
            reader.read_file(non_existent)

        assert "not found" in str(exc_info.value).lower()
        assert "nonexistent" in str(exc_info.value)

    def test_read_file_invalid_format(self):
        """Test reading a non-.xlsx file raises ExcelIOError."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            tmp.write(b"not an excel file")

        try:
            with pytest.raises(ExcelIOError) as exc_info:
                reader.read_file(tmp_path)

            assert "invalid file format" in str(exc_info.value).lower()
            assert ".xlsx" in str(exc_info.value)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_file_empty_worksheet(self):
        """Test reading an empty worksheet returns empty DataFrame."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create empty workbook
            wb = Workbook()
            wb.active.title = "EmptySheet"
            wb.save(tmp_path)
            wb.close()

            # Read the file
            df = reader.read_file(tmp_path)

            # Should return empty DataFrame
            assert isinstance(df, pd.DataFrame)
            assert df.empty

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_sheet_specific_sheet(self):
        """Test reading a specific worksheet by name."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with multiple sheets
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1["A1"] = "Data1"
            ws1["A2"] = "Value1"

            ws2 = wb.create_sheet("Sheet2")
            ws2["A1"] = "Data2"
            ws2["A2"] = "Value2"

            wb.save(tmp_path)
            wb.close()

            # Read specific sheet
            df = reader.read_sheet(tmp_path, "Sheet2")

            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            assert df.columns.tolist() == ["Data2"]
            assert df.iloc[0]["Data2"] == "Value2"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_sheet_nonexistent_sheet(self):
        """Test reading a non-existent sheet raises ExcelIOError."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            wb.active.title = "Sheet1"
            wb.save(tmp_path)
            wb.close()

            with pytest.raises(ExcelIOError) as exc_info:
                reader.read_sheet(tmp_path, "NonExistentSheet")

            assert "not found" in str(exc_info.value).lower()
            assert "NonExistentSheet" in str(exc_info.value)
            assert "Sheet1" in str(exc_info.value)  # Should list available sheets

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_list_sheets(self):
        """Test listing all worksheet names."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with multiple sheets
            wb = Workbook()
            wb.active.title = "FirstSheet"
            wb.create_sheet("SecondSheet")
            wb.create_sheet("ThirdSheet")
            wb.save(tmp_path)
            wb.close()

            # List sheets
            sheets = reader.list_sheets(tmp_path)

            assert isinstance(sheets, list)
            assert len(sheets) == 3
            assert "FirstSheet" in sheets
            assert "SecondSheet" in sheets
            assert "ThirdSheet" in sheets

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_list_sheets_missing_file(self):
        """Test listing sheets from non-existent file raises ExcelIOError."""
        reader = ExcelReader()
        non_existent = Path("/nonexistent/file.xlsx")

        with pytest.raises(ExcelIOError) as exc_info:
            reader.list_sheets(non_existent)

        assert "not found" in str(exc_info.value).lower()

    def test_read_file_directory_instead_of_file(self):
        """Test that providing a directory path raises ExcelIOError."""
        reader = ExcelReader()

        with tempfile.TemporaryDirectory() as tmpdir:
            with pytest.raises(ExcelIOError) as exc_info:
                reader.read_file(Path(tmpdir))

            assert "directory" in str(exc_info.value).lower()

    def test_read_file_with_custom_header(self):
        """Test reading file with custom header row."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with header on row 2
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Skip this"
            ws["A2"] = "Name"
            ws["B2"] = "Amount"
            ws["A3"] = "Entry1"
            ws["B3"] = 100
            wb.save(tmp_path)
            wb.close()

            # Read with header=1 (second row)
            df = reader.read_file(tmp_path, header=1)

            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            assert df.columns.tolist() == ["Name", "Amount"]
            assert df.iloc[0]["Name"] == "Entry1"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_file_multiple_data_types(self):
        """Test reading file with various data types."""
        reader = ExcelReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Text"
            ws["B1"] = "Number"
            ws["C1"] = "Date"
            ws["A2"] = "Test"
            ws["B2"] = 123.45
            ws["C2"] = "2024-01-15"
            wb.save(tmp_path)
            wb.close()

            df = reader.read_file(tmp_path)

            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            assert df.iloc[0]["Text"] == "Test"
            assert df.iloc[0]["Number"] == 123.45

        finally:
            if tmp_path.exists():
                tmp_path.unlink()
