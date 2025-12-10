"""Unit tests for JournalEntryReader class."""

import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from veritas_accounting.excel.journal_reader import JournalEntryReader
from veritas_accounting.models.journal import JournalEntry
from veritas_accounting.utils.exceptions import ExcelIOError, ValidationError


class TestJournalEntryReader:
    """Test suite for JournalEntryReader class."""

    def test_read_journal_entries_basic(self):
        """Test reading basic journal entries from Excel file."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create workbook with journal entries
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Test Entry"
            ws["D2"] = "OL"
            ws["E2"] = 1000.50
            ws["F2"] = "2024-01-15"

            ws["A3"] = "JE-002"
            ws["B3"] = 2024
            ws["C3"] = "Another Entry"
            ws["D3"] = "CR"
            ws["E3"] = 2000.75
            ws["F3"] = "2024-02-20"

            wb.save(tmp_path)
            wb.close()

            # Read entries
            entries, errors = reader.read_journal_entries(tmp_path)

            # Verify results
            assert len(entries) == 2
            assert len(errors) == 0

            entry1 = entries[0]
            assert isinstance(entry1, JournalEntry)
            assert entry1.entry_id == "JE-001"
            assert entry1.year == 2024
            assert entry1.description == "Test Entry"
            assert entry1.old_type == "OL"
            assert entry1.amount == Decimal("1000.50")
            assert entry1.date == datetime(2024, 1, 15)

            entry2 = entries[1]
            assert entry2.entry_id == "JE-002"
            assert entry2.amount == Decimal("2000.75")

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_with_chinese_text(self):
        """Test reading journal entries with Chinese text."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "编号"
            ws["B1"] = "年份"
            ws["C1"] = "描述"
            ws["D1"] = "类型"
            ws["E1"] = "金额"
            ws["F1"] = "日期"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "测试条目"
            ws["D2"] = "OL"
            ws["E2"] = 1500.25
            ws["F2"] = "2024-03-10"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 1
            assert len(errors) == 0
            assert entries[0].description == "测试条目"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_missing_required_column(self):
        """Test reading entries with missing required column raises ValidationError."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            # Missing description, old_type, amount, date
            wb.save(tmp_path)
            wb.close()

            with pytest.raises(ValidationError) as exc_info:
                reader.read_journal_entries(tmp_path)

            assert "Missing required columns" in str(exc_info.value)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_empty_rows(self):
        """Test that empty rows are skipped."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"

            # Row 3 is empty (should be skipped)
            # Row 4 has data
            ws["A4"] = "JE-002"
            ws["B4"] = 2024
            ws["C4"] = "Entry 2"
            ws["D4"] = "CR"
            ws["E4"] = 2000
            ws["F4"] = "2024-02-20"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 2
            assert len(errors) == 0

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_invalid_date(self):
        """Test that invalid dates are caught and reported as errors."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "invalid-date"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 0
            assert len(errors) == 1
            assert "Invalid date format" in str(errors[0])

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_missing_required_field(self):
        """Test that missing required fields in rows are caught."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"

            ws["A3"] = "JE-002"
            ws["B3"] = 2024
            # Missing description, old_type, amount, date
            ws["C3"] = None
            ws["D3"] = None
            ws["E3"] = None
            ws["F3"] = None

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 1  # First entry is valid
            assert len(errors) >= 1  # Second entry has errors
            assert any("Missing required field" in str(e) for e in errors)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_optional_fields(self):
        """Test that optional fields (quarter, notes) are handled correctly."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"
            ws["G1"] = "quarter"
            ws["H1"] = "notes"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"
            ws["G2"] = 1
            ws["H2"] = "Q1 entry"

            ws["A3"] = "JE-002"
            ws["B3"] = 2024
            ws["C3"] = "Entry 2"
            ws["D3"] = "CR"
            ws["E3"] = 2000
            ws["F3"] = "2024-04-20"
            # No quarter or notes (should be None)

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 2
            assert len(errors) == 0

            assert entries[0].quarter == 1
            assert entries[0].notes == "Q1 entry"
            assert entries[1].quarter is None
            assert entries[1].notes is None

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_different_date_formats(self):
        """Test that different date formats are parsed correctly."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"  # YYYY-MM-DD

            ws["A3"] = "JE-002"
            ws["B3"] = 2024
            ws["C3"] = "Entry 2"
            ws["D3"] = "CR"
            ws["E3"] = 2000
            ws["F3"] = "2024/02/20"  # YYYY/MM/DD

            ws["A4"] = "JE-003"
            ws["B4"] = 2024
            ws["C4"] = "Entry 3"
            ws["D4"] = "OL"
            ws["E4"] = 3000
            ws["F4"] = "15/03/2024"  # DD/MM/YYYY

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 3
            assert len(errors) == 0

            assert entries[0].date == datetime(2024, 1, 15)
            assert entries[1].date == datetime(2024, 2, 20)
            assert entries[2].date == datetime(2024, 3, 15)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_custom_column_mapping(self):
        """Test reading with custom column mapping."""
        custom_mapping = {
            "entry_id": ["ID"],
            "year": ["YEAR"],
            "description": ["DESC"],
            "old_type": ["TYPE"],
            "amount": ["AMT"],
            "date": ["DATE"],
        }
        reader = JournalEntryReader(column_mapping=custom_mapping)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "ID"
            ws["B1"] = "YEAR"
            ws["C1"] = "DESC"
            ws["D1"] = "TYPE"
            ws["E1"] = "AMT"
            ws["F1"] = "DATE"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Test Entry"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 1
            assert len(errors) == 0
            assert entries[0].entry_id == "JE-001"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_specific_sheet(self):
        """Test reading from a specific worksheet."""
        reader = JournalEntryReader(sheet_name="JournalEntries")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "OtherSheet"
            ws1["A1"] = "Other Data"

            ws2 = wb.create_sheet("JournalEntries")
            ws2["A1"] = "entry_id"
            ws2["B1"] = "year"
            ws2["C1"] = "description"
            ws2["D1"] = "old_type"
            ws2["E1"] = "amount"
            ws2["F1"] = "date"

            ws2["A2"] = "JE-001"
            ws2["B2"] = 2024
            ws2["C2"] = "Test Entry"
            ws2["D2"] = "OL"
            ws2["E2"] = 1000
            ws2["F2"] = "2024-01-15"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 1
            assert len(errors) == 0
            assert entries[0].entry_id == "JE-001"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_invalid_amount(self):
        """Test that invalid amounts are caught and reported."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Entry 1"
            ws["D2"] = "OL"
            ws["E2"] = "not-a-number"
            ws["F2"] = "2024-01-15"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 0
            assert len(errors) >= 1
            assert any("Invalid amount" in str(e) for e in errors)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_collects_all_errors(self):
        """Test that all validation errors are collected (not fail-fast)."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "entry_id"
            ws["B1"] = "year"
            ws["C1"] = "description"
            ws["D1"] = "old_type"
            ws["E1"] = "amount"
            ws["F1"] = "date"

            # Row 2: Valid entry
            ws["A2"] = "JE-001"
            ws["B2"] = 2024
            ws["C2"] = "Valid Entry"
            ws["D2"] = "OL"
            ws["E2"] = 1000
            ws["F2"] = "2024-01-15"

            # Row 3: Invalid date
            ws["A3"] = "JE-002"
            ws["B3"] = 2024
            ws["C3"] = "Invalid Date"
            ws["D3"] = "OL"
            ws["E3"] = 2000
            ws["F3"] = "invalid-date"

            # Row 4: Invalid amount
            ws["A4"] = "JE-003"
            ws["B4"] = 2024
            ws["C4"] = "Invalid Amount"
            ws["D4"] = "OL"
            ws["E4"] = "not-a-number"
            ws["F4"] = "2024-03-15"

            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            # Should have 1 valid entry and 2 errors
            assert len(entries) == 1
            assert len(errors) == 2
            assert entries[0].entry_id == "JE-001"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_empty_file(self):
        """Test reading an empty Excel file."""
        reader = JournalEntryReader()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            wb.save(tmp_path)
            wb.close()

            entries, errors = reader.read_journal_entries(tmp_path)

            assert len(entries) == 0
            assert len(errors) == 0

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_read_journal_entries_missing_file(self):
        """Test that missing file raises ExcelIOError."""
        reader = JournalEntryReader()
        non_existent = Path("/nonexistent/journal.xlsx")

        with pytest.raises(ExcelIOError):
            reader.read_journal_entries(non_existent)
