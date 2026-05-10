"""Tests for LedgerOutputGenerator categorization sheet name and sort order."""

from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook

from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.reporting.ledger_output import LedgerOutputGenerator


def _le(**kwargs: object) -> LedgerEntry:
    base = dict(
        entry_id="LE-x",
        account_code="1100",
        account_path="银行存款",
        amount=Decimal("10"),
        date=datetime(2024, 2, 15),
        description="d",
        source_entry_id="X-1-1",
        rule_applied="R",
        quarter=1,
        year=2024,
        ledger_type="DR",
    )
    base.update(kwargs)
    return LedgerEntry(**base)


class TestJournalCategorizationSheet:
    def test_sheet_name_and_sort_order(self) -> None:
        """Same quarter/date: order A, I, B, then others; earlier dates first."""
        entries = [
            _le(entry_id="LE-2", source_entry_id="B-1-1", date=datetime(2024, 3, 10)),
            _le(entry_id="LE-1", source_entry_id="A-1-2", date=datetime(2024, 3, 10)),
            _le(entry_id="LE-4", source_entry_id="W-1-1", date=datetime(2024, 3, 10)),
            _le(entry_id="LE-3", source_entry_id="I-1-1", date=datetime(2024, 3, 10)),
            _le(entry_id="LE-0", source_entry_id="A-1-1", date=datetime(2024, 3, 5)),
        ]
        gen = LedgerOutputGenerator(None)
        wb = Workbook()
        wb.remove(wb.active)
        gen.add_ledger_entries_sheet(wb, entries, index=0)

        assert wb.sheetnames == [LedgerOutputGenerator.JOURNAL_CATEGORIZATION_SHEET]
        ws = wb[LedgerOutputGenerator.JOURNAL_CATEGORIZATION_SHEET]
        # Column 12 = Source Entry ID
        ids = [ws.cell(row=r, column=12).value for r in range(2, 7)]
        assert ids == ["A-1-1", "A-1-2", "I-1-1", "B-1-1", "W-1-1"]

    def test_prefix_rank_constants(self) -> None:
        assert LedgerOutputGenerator._source_entry_prefix_rank("A-3-1") == 0
        assert LedgerOutputGenerator._source_entry_prefix_rank("I-2-1") == 1
        assert LedgerOutputGenerator._source_entry_prefix_rank("B-3-11") == 2
        assert LedgerOutputGenerator._source_entry_prefix_rank("W-1-18") == 3
