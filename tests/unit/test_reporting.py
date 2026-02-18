"""Unit tests for reporting modules."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest

from veritas_accounting.models.account import Account, AccountHierarchy
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.reporting.ledger_output import LedgerOutputGenerator
from veritas_accounting.reporting.quarterly_report import QuarterlyReportGenerator
from veritas_accounting.reporting.formatting import ExcelFormatter
from veritas_accounting.utils.exceptions import ExcelIOError


@pytest.fixture
def sample_account_hierarchy():
    """Create a sample account hierarchy for testing."""
    accounts = [
        Account(
            code="A1",
            name="Level 1 Account 1",
            level=1,
            parent_code=None,
            full_path="Level1/Account1",
        ),
        Account(
            code="A1-1",
            name="Level 2 Account 1-1",
            level=2,
            parent_code="A1",
            full_path="Level1/Account1/Level2/Account1-1",
        ),
        Account(
            code="A1-1-1",
            name="Level 3 Account 1-1-1",
            level=3,
            parent_code="A1-1",
            full_path="Level1/Account1/Level2/Account1-1/Level3/Account1-1-1",
        ),
    ]
    return AccountHierarchy(accounts)


@pytest.fixture
def ledger_new_style_hierarchy():
    """Create account hierarchy in ledger_new style (4-digit ledger IDs)."""
    accounts = [
        Account(
            code="1100",  # 4-digit ledger ID
            name="银行存款",  # Account name (used in mapping rules)
            level=1,
            parent_code=None,
            full_path="资产 > 银行存款",
        ),
        Account(
            code="4301",  # 4-digit ledger ID
            name="收入OL",  # Account name (used in mapping rules)
            level=1,
            parent_code=None,
            full_path="收入 > 收入OL",
        ),
        Account(
            code="5301",  # 4-digit ledger ID
            name="支出OL讲师",  # Account name (used in mapping rules)
            level=1,
            parent_code=None,
            full_path="支出 > 支出OL讲师",
        ),
    ]
    return AccountHierarchy(accounts)


@pytest.fixture
def sample_ledger_entries():
    """Create sample ledger entries for testing."""
    return [
        LedgerEntry(
            entry_id="LE-001",
            account_code="A1",
            account_path="Level1/Account1",
            amount=Decimal("1000.00"),
            date=datetime(2024, 1, 15),
            description="Test entry 1",
            source_entry_id="JE-001",
            rule_applied="R-001",
            quarter=1,
            year=2024,
        ),
        LedgerEntry(
            entry_id="LE-002",
            account_code="A1-1",
            account_path="Level1/Account1/Level2/Account1-1",
            amount=Decimal("500.00"),
            date=datetime(2024, 2, 20),
            description="Test entry 2",
            source_entry_id="JE-002",
            rule_applied="R-002",
            quarter=1,
            year=2024,
        ),
        LedgerEntry(
            entry_id="LE-003",
            account_code="A1-1-1",
            account_path="Level1/Account1/Level2/Account1-1/Level3/Account1-1-1",
            amount=Decimal("250.00"),
            date=datetime(2024, 3, 10),
            description="Test entry 3",
            source_entry_id="JE-003",
            rule_applied="R-003",
            quarter=1,
            year=2024,
        ),
    ]


class TestLedgerOutputGenerator:
    """Test cases for LedgerOutputGenerator class."""

    def test_initialization(self):
        """Test ledger output generator initialization."""
        generator = LedgerOutputGenerator()
        assert generator.account_hierarchy is None

    def test_initialization_with_hierarchy(self, sample_account_hierarchy):
        """Test initialization with account hierarchy."""
        generator = LedgerOutputGenerator(sample_account_hierarchy)
        assert generator.account_hierarchy == sample_account_hierarchy

    def test_generate_output(
        self, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test generating ledger output Excel file."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "ledger_output.xlsx"
            generator = LedgerOutputGenerator(sample_account_hierarchy)
            generator.generate(sample_ledger_entries, output_path)

            assert output_path.exists()
            assert output_path.suffix == ".xlsx"

    def test_generate_output_without_hierarchy(self, sample_ledger_entries):
        """Test generating output without account hierarchy."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "ledger_output.xlsx"
            generator = LedgerOutputGenerator()
            generator.generate(sample_ledger_entries, output_path)

            assert output_path.exists()

    def test_organize_by_hierarchy(
        self, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test organizing entries by hierarchy."""
        generator = LedgerOutputGenerator(sample_account_hierarchy)
        organized = generator._organize_by_hierarchy(sample_ledger_entries)

        assert len(organized) == len(sample_ledger_entries)
        # Should be organized by level (Level 1 → Level 3)
        assert organized[0].account_code == "A1"
        assert organized[1].account_code == "A1-1"
        assert organized[2].account_code == "A1-1-1"

    def test_calculate_level_totals(
        self, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test calculating level totals."""
        generator = LedgerOutputGenerator(sample_account_hierarchy)
        level_totals = generator._calculate_level_totals(sample_ledger_entries)

        assert 1 in level_totals
        assert 2 in level_totals
        assert 3 in level_totals
        assert level_totals[1]["A1"]["amount"] == 1000.0
        assert level_totals[2]["A1-1"]["amount"] == 500.0
        assert level_totals[3]["A1-1-1"]["amount"] == 250.0

    def test_ledger_id_in_output_by_code(self, ledger_new_style_hierarchy, tmp_path):
        """Test that ledger ID (4-digit number) is included in output when account_code matches."""
        generator = LedgerOutputGenerator(ledger_new_style_hierarchy)
        
        # Create ledger entry with account_code matching hierarchy code
        ledger_entries = [
            LedgerEntry(
                entry_id="LE-001",
                account_code="1100",  # Matches hierarchy code
                account_path="资产 > 银行存款",
                amount=Decimal("1000.00"),
                date=datetime(2024, 1, 15),
                description="Test entry",
                source_entry_id="JE-001",
                rule_applied="R-001",
                quarter=1,
                year=2024,
                ledger_type="CR",
            ),
        ]
        
        output_path = tmp_path / "test_ledger_id.xlsx"
        generator.generate(ledger_entries, output_path)
        
        # Read back and verify ledger ID is in output
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Ledger Entries"]
        
        # Check header row
        assert ws.cell(row=1, column=2).value == "Ledger ID"
        # Check data row - ledger ID should be "1100" (4-digit number)
        assert ws.cell(row=2, column=2).value == "1100"
        # Account Code should also be "1100" in this case
        assert ws.cell(row=2, column=3).value == "1100"

    def test_ledger_id_in_output_by_name(self, ledger_new_style_hierarchy, tmp_path):
        """Test that ledger ID is found by account name when mapping rules use names."""
        generator = LedgerOutputGenerator(ledger_new_style_hierarchy)
        
        # Create ledger entry with account_code matching hierarchy name (as mapping rules do)
        ledger_entries = [
            LedgerEntry(
                entry_id="LE-001",
                account_code="银行存款",  # Account name (as used in mapping rules)
                account_path="资产 > 银行存款",
                amount=Decimal("1000.00"),
                date=datetime(2024, 1, 15),
                description="Test entry",
                source_entry_id="JE-001",
                rule_applied="R-001",
                quarter=1,
                year=2024,
                ledger_type="CR",
            ),
        ]
        
        output_path = tmp_path / "test_ledger_id_by_name.xlsx"
        generator.generate(ledger_entries, output_path)
        
        # Read back and verify ledger ID is found by name lookup
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Ledger Entries"]
        
        # Check data row - ledger ID should be "1100" (found by name lookup)
        assert ws.cell(row=2, column=2).value == "1100"
        # Account Code should be "银行存款" (original from mapping rules)
        assert ws.cell(row=2, column=3).value == "银行存款"

    def test_get_account_level(
        self, sample_account_hierarchy
    ):
        """Test getting account level."""
        generator = LedgerOutputGenerator(sample_account_hierarchy)
        assert generator._get_account_level("A1") == 1
        assert generator._get_account_level("A1-1") == 2
        assert generator._get_account_level("A1-1-1") == 3
        assert generator._get_account_level("NONEXISTENT") == 0


class TestQuarterlyReportGenerator:
    """Test cases for QuarterlyReportGenerator class."""

    def test_initialization(self):
        """Test quarterly report generator initialization."""
        generator = QuarterlyReportGenerator()
        assert generator.account_hierarchy is None
        assert generator.aggregator is not None

    def test_initialization_with_hierarchy(self, sample_account_hierarchy):
        """Test initialization with account hierarchy."""
        generator = QuarterlyReportGenerator(sample_account_hierarchy)
        assert generator.account_hierarchy == sample_account_hierarchy

    def test_generate_report(
        self, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test generating quarterly report Excel file."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "quarterly_report.xlsx"
            generator = QuarterlyReportGenerator(sample_account_hierarchy)
            generator.generate(sample_ledger_entries, output_path)

            assert output_path.exists()
            assert output_path.suffix == ".xlsx"

    def test_ledger_id_in_quarterly_report(self, ledger_new_style_hierarchy, tmp_path):
        """Test that ledger ID is included in quarterly report output."""
        generator = QuarterlyReportGenerator(ledger_new_style_hierarchy)
        
        ledger_entries = [
            LedgerEntry(
                entry_id="LE-001",
                account_code="收入OL",  # Account name (as used in mapping rules)
                account_path="收入 > 收入OL",
                amount=Decimal("5000.00"),
                date=datetime(2024, 1, 15),
                description="Test entry",
                source_entry_id="JE-001",
                rule_applied="R-001",
                quarter=1,
                year=2024,
                ledger_type="CR",
            ),
        ]
        
        output_path = tmp_path / "test_quarterly_ledger_id.xlsx"
        generator.generate(ledger_entries, output_path)
        
        # Read back and verify ledger ID is in quarterly report
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Quarterly Totals"]
        
        # Check header row
        assert ws.cell(row=3, column=1).value == "Ledger ID"
        # Check data row - ledger ID should be "4301" (found by name lookup)
        assert ws.cell(row=5, column=1).value == "4301"
        # Account Code should be "收入OL" (original from mapping rules)
        assert ws.cell(row=5, column=2).value == "收入OL"

    def test_generate_report_without_charts(
        self, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test generating report without charts."""
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "quarterly_report.xlsx"
            generator = QuarterlyReportGenerator(sample_account_hierarchy)
            generator.generate(sample_ledger_entries, output_path, include_charts=False)

            assert output_path.exists()

    def test_get_quarter_color(self):
        """Test getting quarter color."""
        generator = QuarterlyReportGenerator()
        assert generator._get_quarter_color(1) == generator.COLOR_QUARTER1
        assert generator._get_quarter_color(2) == generator.COLOR_QUARTER2
        assert generator._get_quarter_color(3) == generator.COLOR_QUARTER3
        assert generator._get_quarter_color(4) == generator.COLOR_QUARTER4
        assert generator._get_quarter_color(5) is None


class TestExcelFormatter:
    """Test cases for ExcelFormatter class."""

    def test_apply_header_style(self):
        """Test applying header style."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.apply_header_style(cell, "Test Header")
        assert cell.value == "Test Header"
        assert cell.font.bold is True
        # openpyxl adds "00" prefix to RGB codes
        assert cell.font.color.rgb in ["FFFFFF", "00FFFFFF"]

    def test_apply_total_style(self):
        """Test applying total style."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.apply_total_style(cell, "Total")
        assert cell.value == "Total"
        assert cell.font.bold is True

    def test_apply_level_style(self):
        """Test applying level style."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.apply_level_style(cell, 1)
        # openpyxl adds "00" prefix to RGB codes
        assert cell.fill.start_color.rgb in [
            ExcelFormatter.COLOR_LEVEL1,
            f"00{ExcelFormatter.COLOR_LEVEL1}",
        ]

    def test_apply_quarter_style(self):
        """Test applying quarter style."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.apply_quarter_style(cell, 1)
        # openpyxl adds "00" prefix to RGB codes
        assert cell.fill.start_color.rgb in [
            ExcelFormatter.COLOR_QUARTER1,
            f"00{ExcelFormatter.COLOR_QUARTER1}",
        ]

    def test_apply_severity_style(self):
        """Test applying severity style."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.apply_severity_style(cell, "error")
        # openpyxl adds "00" prefix to RGB codes
        assert cell.fill.start_color.rgb in [
            ExcelFormatter.COLOR_ERROR,
            f"00{ExcelFormatter.COLOR_ERROR}",
        ]

    def test_format_amount_cell(self):
        """Test formatting amount cell."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.format_amount_cell(cell, 1234.56)
        assert cell.value == 1234.56
        assert cell.number_format == "#,##0.00"

    def test_format_percentage_cell(self):
        """Test formatting percentage cell."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.format_percentage_cell(cell, 0.25)
        assert cell.value == 0.25
        assert cell.number_format == "0.00%"

    def test_format_date_cell(self):
        """Test formatting date cell."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]

        ExcelFormatter.format_date_cell(cell, "2024-01-15")
        assert cell.value == "2024-01-15"
        assert cell.number_format == "YYYY-MM-DD"

    def test_auto_adjust_column_widths(self):
        """Test auto-adjusting column widths."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Short"
        ws["B1"] = "This is a much longer text"

        ExcelFormatter.auto_adjust_column_widths(ws)
        assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width

    def test_setup_filters(self):
        """Test setting up filters."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = "Data"

        ExcelFormatter.setup_filters(ws)
        assert ws.auto_filter.ref is not None

    def test_freeze_panes(self):
        """Test freezing panes."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ExcelFormatter.freeze_panes(ws, "A2")
        assert ws.freeze_panes == "A2"








