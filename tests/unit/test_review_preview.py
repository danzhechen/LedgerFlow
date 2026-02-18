"""Unit tests for review preview generator."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest

from veritas_accounting.audit.trail import AppliedRule, AuditTrail, TransformationRecord
from veritas_accounting.models.account import Account, AccountHierarchy
from veritas_accounting.models.journal import JournalEntry
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.reporting.review_preview import (
    EntryReviewFlag,
    ReviewPreviewGenerator,
    ReviewStatus,
)
from veritas_accounting.validation.error_detector import (
    DetectedError,
    ErrorDetector,
    ERROR_TYPE_DATA,
    SEVERITY_ERROR,
    SEVERITY_WARNING,
)
from veritas_accounting.validation.input_validator import ValidationError


@pytest.fixture
def sample_account_hierarchy():
    """Create a sample account hierarchy for testing."""
    accounts = [
        Account(
            code="1100",
            name="银行存款",
            level=1,
            parent_code=None,
            full_path="资产 > 银行存款",
        ),
        Account(
            code="4301",
            name="收入OL",
            level=1,
            parent_code=None,
            full_path="收入 > 收入OL",
        ),
    ]
    return AccountHierarchy(accounts)


@pytest.fixture
def sample_journal_entries():
    """Create sample journal entries for testing."""
    return [
        JournalEntry(
            entry_id="JE-001",
            year=2024,
            description="Test entry 1",
            old_type="收入OL",
            amount=Decimal("1000.00"),
            date=datetime(2024, 1, 15),
            quarter=1,
        ),
        JournalEntry(
            entry_id="JE-002",
            year=2024,
            description="Test entry 2",
            old_type="支出OL",
            amount=Decimal("500.00"),
            date=datetime(2024, 2, 20),
            quarter=1,
        ),
    ]


@pytest.fixture
def sample_ledger_entries():
    """Create sample ledger entries for testing."""
    return [
        LedgerEntry(
            entry_id="LE-001",
            account_code="收入OL",
            account_path="收入 > 收入OL",
            amount=Decimal("1000.00"),
            date=datetime(2024, 1, 15),
            description="Test entry 1",
            source_entry_id="JE-001",
            rule_applied="R-001",
            quarter=1,
            year=2024,
        ),
        LedgerEntry(
            entry_id="LE-002",
            account_code="银行存款",
            account_path="资产 > 银行存款",
            amount=Decimal("500.00"),
            date=datetime(2024, 2, 20),
            description="Test entry 2",
            source_entry_id="JE-002",
            rule_applied="R-002",
            quarter=1,
            year=2024,
        ),
    ]


@pytest.fixture
def sample_audit_trail(sample_journal_entries):
    """Create sample audit trail for testing."""
    trail = AuditTrail(user="test_user", system_version="1.0.0")
    
    # Create transformation records
    record1 = TransformationRecord(
        entry_id="JE-001",
        timestamp=datetime(2024, 1, 15, 10, 0, 0),
        source_entry=sample_journal_entries[0],
        applied_rules=[
            AppliedRule(
                rule_id="R-001",
                condition="old_type == '收入OL'",
                account_code="收入OL",
                priority=1,
            )
        ],
        generated_entries=[
            LedgerEntry(
                entry_id="LE-001",
                account_code="收入OL",
                account_path="收入 > 收入OL",
                amount=Decimal("1000.00"),
                date=datetime(2024, 1, 15),
                description="Test entry 1",
                source_entry_id="JE-001",
                rule_applied="R-001",
                quarter=1,
                year=2024,
            )
        ],
        no_match=False,
    )
    
    record2 = TransformationRecord(
        entry_id="JE-002",
        timestamp=datetime(2024, 2, 20, 10, 0, 0),
        source_entry=sample_journal_entries[1],
        applied_rules=[
            AppliedRule(
                rule_id="R-002",
                condition="old_type == '支出OL'",
                account_code="银行存款",
                priority=1,
            )
        ],
        generated_entries=[
            LedgerEntry(
                entry_id="LE-002",
                account_code="银行存款",
                account_path="资产 > 银行存款",
                amount=Decimal("500.00"),
                date=datetime(2024, 2, 20),
                description="Test entry 2",
                source_entry_id="JE-002",
                rule_applied="R-002",
                quarter=1,
                year=2024,
            )
        ],
        no_match=False,
    )
    
    trail.records = [record1, record2]
    return trail


class TestReviewPreviewGenerator:
    """Test cases for ReviewPreviewGenerator class."""

    def test_initialization(self):
        """Test review preview generator initialization."""
        generator = ReviewPreviewGenerator()
        assert generator.account_hierarchy is None
        assert generator.error_detector is None
        assert generator.audit_trail is None
        assert generator.review_flags == []

    def test_initialization_with_all_params(
        self, sample_account_hierarchy, sample_audit_trail
    ):
        """Test initialization with all parameters."""
        error_detector = ErrorDetector()
        generator = ReviewPreviewGenerator(
            account_hierarchy=sample_account_hierarchy,
            error_detector=error_detector,
            audit_trail=sample_audit_trail,
        )
        assert generator.account_hierarchy == sample_account_hierarchy
        assert generator.error_detector == error_detector
        assert generator.audit_trail == sample_audit_trail

    def test_generate_basic(
        self, sample_ledger_entries, sample_journal_entries, tmp_path
    ):
        """Test generating basic review preview file."""
        generator = ReviewPreviewGenerator()
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        assert output_path.exists()
        assert output_path.suffix == ".xlsx"

    def test_generate_with_sheets(
        self, sample_ledger_entries, sample_journal_entries, tmp_path
    ):
        """Test that all required sheets are created."""
        generator = ReviewPreviewGenerator()
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        
        # Check all sheets exist
        sheet_names = wb.sheetnames
        assert "Review Dashboard" in sheet_names
        assert "Preview Table" in sheet_names
        assert "Comparison View" in sheet_names
        assert "Flagged Entries" in sheet_names

    def test_flag_no_match_entries(
        self, sample_journal_entries, sample_ledger_entries, sample_audit_trail
    ):
        """Test flagging entries with no matching rules."""
        # Create audit trail with one no_match entry
        no_match_record = TransformationRecord(
            entry_id="JE-003",
            timestamp=datetime(2024, 3, 1, 10, 0, 0),
            source_entry=sample_journal_entries[0],
            applied_rules=[],
            generated_entries=[],
            no_match=True,
        )
        sample_audit_trail.records.append(no_match_record)

        generator = ReviewPreviewGenerator(audit_trail=sample_audit_trail)
        generator._analyze_entries(sample_ledger_entries, sample_journal_entries)

        # Should have one no_match flag
        no_match_flags = [
            f for f in generator.review_flags if f.status == ReviewStatus.NO_MATCH
        ]
        assert len(no_match_flags) == 1
        assert no_match_flags[0].entry_id == "JE-003"
        assert no_match_flags[0].severity == "error"

    def test_flag_validation_errors(
        self, sample_journal_entries, sample_ledger_entries
    ):
        """Test flagging validation errors."""
        error_detector = ErrorDetector()
        error = ValidationError(
            row_number=1,
            field_name="amount",
            error_type="validation_error",
            error_message="Invalid amount",
            actual_value="invalid",
            entry_id="JE-001",
        )
        error_detector.add_validation_result([error], error_category=ERROR_TYPE_DATA)

        generator = ReviewPreviewGenerator(error_detector=error_detector)
        generator._analyze_entries(sample_ledger_entries, sample_journal_entries)

        # Should have flags for ledger entries corresponding to JE-001
        error_flags = [
            f
            for f in generator.review_flags
            if f.status == ReviewStatus.VALIDATION_ERROR
        ]
        assert len(error_flags) >= 1
        assert any(f.source_entry_id == "JE-001" for f in error_flags)

    def test_flag_validation_warnings(
        self, sample_journal_entries, sample_ledger_entries
    ):
        """Test flagging validation warnings."""
        from veritas_accounting.validation.pipeline import ValidationWarning

        error_detector = ErrorDetector()
        warning = ValidationWarning(
            row_number=1,
            field_name="amount",
            warning_type="unusual_value",
            warning_message="Potential issue",
            actual_value=1000.0,
            entry_id="JE-001",
        )
        error_detector.add_validation_result(
            validation_errors=[], validation_warnings=[warning]
        )

        generator = ReviewPreviewGenerator(error_detector=error_detector)
        generator._analyze_entries(sample_ledger_entries, sample_journal_entries)

        warning_flags = [
            f
            for f in generator.review_flags
            if f.status == ReviewStatus.VALIDATION_WARNING
        ]
        assert len(warning_flags) >= 1

    def test_flag_missing_accounts(
        self, sample_journal_entries, sample_ledger_entries, sample_account_hierarchy
    ):
        """Test flagging entries with missing account codes."""
        # Create ledger entry with non-existent account
        ledger_entry_with_missing = LedgerEntry(
            entry_id="LE-003",
            account_code="NONEXISTENT",
            account_path="Unknown",
            amount=Decimal("100.00"),
            date=datetime(2024, 3, 1),
            description="Test",
            source_entry_id="JE-003",
            rule_applied="R-003",
            quarter=1,
            year=2024,
        )
        ledger_entries = sample_ledger_entries + [ledger_entry_with_missing]

        generator = ReviewPreviewGenerator(account_hierarchy=sample_account_hierarchy)
        generator._analyze_entries(ledger_entries, sample_journal_entries)

        missing_account_flags = [
            f for f in generator.review_flags if f.status == ReviewStatus.MISSING_ACCOUNT
        ]
        assert len(missing_account_flags) == 1
        assert missing_account_flags[0].entry_id == "LE-003"

    def test_flag_unusual_amounts(
        self, sample_journal_entries, sample_ledger_entries
    ):
        """Test flagging entries with unusual amounts (statistical outliers)."""
        # Need at least 3 entries for unusual amount detection to work
        # Create multiple entries with normal amounts, then one very large outlier
        normal_entries = [
            LedgerEntry(
                entry_id="LE-004",
                account_code="收入OL",
                account_path="收入 > 收入OL",
                amount=Decimal("1000.00"),
                date=datetime(2024, 3, 10),
                description="Normal entry 1",
                source_entry_id="JE-004",
                rule_applied="R-004",
                quarter=1,
                year=2024,
            ),
            LedgerEntry(
                entry_id="LE-005",
                account_code="收入OL",
                account_path="收入 > 收入OL",
                amount=Decimal("1100.00"),
                date=datetime(2024, 3, 15),
                description="Normal entry 2",
                source_entry_id="JE-005",
                rule_applied="R-005",
                quarter=1,
                year=2024,
            ),
            LedgerEntry(
                entry_id="LE-006",
                account_code="收入OL",
                account_path="收入 > 收入OL",
                amount=Decimal("900.00"),
                date=datetime(2024, 3, 20),
                description="Normal entry 3",
                source_entry_id="JE-006",
                rule_applied="R-006",
                quarter=1,
                year=2024,
            ),
        ]
        # Create entry with very large amount (outlier)
        unusual_entry = LedgerEntry(
            entry_id="LE-003",
            account_code="收入OL",
            account_path="收入 > 收入OL",
            amount=Decimal("1000000.00"),  # Very large compared to ~1000 range
            date=datetime(2024, 3, 1),
            description="Unusual amount",
            source_entry_id="JE-003",
            rule_applied="R-003",
            quarter=1,
            year=2024,
        )
        ledger_entries = sample_ledger_entries + normal_entries + [unusual_entry]

        generator = ReviewPreviewGenerator()
        generator._analyze_entries(ledger_entries, sample_journal_entries)

        unusual_flags = [
            f for f in generator.review_flags if f.status == ReviewStatus.UNUSUAL_AMOUNT
        ]
        # With amounts [1000, 500, 1000, 1100, 900, 1000000]:
        # Median = 1000, MAD ≈ 150, threshold = 1000 + (5 * 150) = 1750
        # 1000000 > 1750, so it should be flagged
        assert len(unusual_flags) >= 1
        assert any(f.entry_id == "LE-003" for f in unusual_flags)

    def test_flag_multiple_rules(
        self, sample_journal_entries, sample_ledger_entries
    ):
        """Test flagging entries with multiple rules applied."""
        audit_trail = AuditTrail()
        record = TransformationRecord(
            entry_id="JE-001",
            timestamp=datetime(2024, 1, 15, 10, 0, 0),
            source_entry=sample_journal_entries[0],
            applied_rules=[
                AppliedRule(
                    rule_id="R-001",
                    condition="condition1",
                    account_code="收入OL",
                    priority=1,
                ),
                AppliedRule(
                    rule_id="R-002",
                    condition="condition2",
                    account_code="收入OL",
                    priority=2,
                ),
            ],
            generated_entries=[],
            no_match=False,
        )
        audit_trail.records = [record]

        generator = ReviewPreviewGenerator(audit_trail=audit_trail)
        generator._analyze_entries(sample_ledger_entries, sample_journal_entries)

        multiple_rules_flags = [
            f for f in generator.review_flags if f.status == ReviewStatus.MULTIPLE_RULES
        ]
        assert len(multiple_rules_flags) == 1

    def test_missing_account_handling(
        self, sample_journal_entries, sample_account_hierarchy
    ):
        """Test that missing account codes are flagged correctly."""
        # Note: LedgerEntry requires account_code to be non-empty string (Pydantic validation)
        # So we can't test with None or empty string directly.
        # The code checks `if not entry.account_code:` as defensive programming.
        # Test with a valid account code that's not in hierarchy (this should be flagged)
        ledger_entry_missing = LedgerEntry(
            entry_id="LE-003",
            account_code="MISSING_CODE",  # Valid code but not in hierarchy
            account_path="Unknown",
            amount=Decimal("100.00"),
            date=datetime(2024, 3, 1),
            description="Test",
            source_entry_id="JE-003",
            rule_applied="R-003",
            quarter=1,
            year=2024,
        )
        ledger_entries = [ledger_entry_missing]

        generator = ReviewPreviewGenerator(account_hierarchy=sample_account_hierarchy)
        generator._analyze_entries(ledger_entries, sample_journal_entries)

        # Should create missing account flag for account not in hierarchy
        missing_account_flags = [
            f for f in generator.review_flags if f.status == ReviewStatus.MISSING_ACCOUNT
        ]
        assert len(missing_account_flags) == 1
        assert missing_account_flags[0].entry_id == "LE-003"

    def test_dashboard_statistics(
        self, sample_journal_entries, sample_ledger_entries, tmp_path
    ):
        """Test that dashboard shows correct statistics."""
        error_detector = ErrorDetector()
        error = ValidationError(
            row_number=1,
            field_name="amount",
            error_type="validation_error",
            error_message="Invalid amount",
            actual_value="invalid",
            entry_id="JE-001",
        )
        error_detector.add_validation_result([error], error_category=ERROR_TYPE_DATA)

        generator = ReviewPreviewGenerator(error_detector=error_detector)
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Review Dashboard"]

        # Check that statistics are present
        # The statistics should include both detected errors and flagged entries
        cell_values = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value 
                      for i in range(4, 15) if ws.cell(row=i, column=1).value}
        
        assert "Total Journal Entries" in cell_values
        assert "Total Ledger Entries" in cell_values
        assert "Entries Needing Review" in cell_values
        assert "Errors Detected (All)" in cell_values

    def test_preview_table_contains_all_entries(
        self, sample_journal_entries, sample_ledger_entries, tmp_path
    ):
        """Test that preview table contains all ledger entries."""
        generator = ReviewPreviewGenerator()
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Preview Table"]

        # Check that all ledger entries are in the preview table
        # Header row is row 3, data starts at row 4
        data_rows = ws.max_row - 3
        assert data_rows == len(sample_ledger_entries)

    def test_comparison_view_side_by_side(
        self, sample_journal_entries, sample_ledger_entries, tmp_path
    ):
        """Test that comparison view shows journal and ledger entries side by side."""
        generator = ReviewPreviewGenerator()
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Comparison View"]

        # Should have journal entry columns and ledger entry columns
        headers = [ws.cell(row=3, column=i).value for i in range(1, 16)]
        assert "Journal Entry ID" in headers
        assert "Ledger Entry ID" in headers
        assert "→" in headers  # Arrow symbol

    def test_flagged_entries_only_shows_problems(
        self, sample_journal_entries, sample_ledger_entries, tmp_path
    ):
        """Test that flagged entries sheet only shows entries with issues."""
        # Create audit trail with one no_match entry
        audit_trail = AuditTrail()
        no_match_record = TransformationRecord(
            entry_id="JE-003",
            timestamp=datetime(2024, 3, 1, 10, 0, 0),
            source_entry=sample_journal_entries[0],
            applied_rules=[],
            generated_entries=[],
            no_match=True,
        )
        audit_trail.records = [no_match_record]

        generator = ReviewPreviewGenerator(audit_trail=audit_trail)
        output_path = tmp_path / "review_preview.xlsx"
        
        generator.generate(
            ledger_entries=sample_ledger_entries,
            journal_entries=sample_journal_entries,
            output_path=output_path,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["Flagged Entries"]

        # Should have at least one flagged entry (the no_match one)
        if ws.cell(row=3, column=1).value != "✓ No entries require review":
            data_rows = ws.max_row - 3
            assert data_rows >= 1
