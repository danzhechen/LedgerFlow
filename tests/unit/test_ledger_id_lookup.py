"""Unit tests specifically for ledger ID lookup functionality."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest

from veritas_accounting.models.account import Account, AccountHierarchy
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.reporting.ledger_output import LedgerOutputGenerator
from veritas_accounting.reporting.quarterly_report import QuarterlyReportGenerator


@pytest.fixture
def real_world_hierarchy():
    """Create account hierarchy matching real-world ledger_new format."""
    accounts = [
        Account(code="1100", name="银行存款", level=1, parent_code=None, full_path="资产 > 银行存款"),
        Account(code="1200", name="其他货币资金", level=1, parent_code=None, full_path="资产 > 其他货币资金"),
        Account(code="4301", name="收入OL", level=1, parent_code=None, full_path="收入 > 收入OL"),
        Account(code="5301", name="支出OL讲师", level=1, parent_code=None, full_path="支出 > 支出OL讲师"),
        Account(code="5101", name="宣传推广", level=1, parent_code=None, full_path="支出 > 宣传推广"),
        Account(code="2209", name="应付OL押金", level=1, parent_code=None, full_path="应付款 > 应付OL押金"),
        Account(code="2201", name="应付OL奖助", level=1, parent_code=None, full_path="应付款 > 应付OL奖助"),
    ]
    return AccountHierarchy(accounts)


def test_ledger_id_lookup_in_ledger_output(real_world_hierarchy, tmp_path):
    """Test that ledger ID (4-digit) is correctly looked up and displayed in ledger output."""
    generator = LedgerOutputGenerator(real_world_hierarchy)
    
    # Create entries with account names (as mapping rules use)
    ledger_entries = [
        LedgerEntry(
            entry_id="LE-001",
            account_code="银行存款",  # Account name from mapping rules
            account_path="资产 > 银行存款",
            amount=Decimal("1000.00"),
            date=datetime(2024, 1, 15),
            description="Test deposit",
            source_entry_id="JE-001",
            rule_applied="R-001",
            quarter=1,
            year=2024,
            ledger_type="CR",
        ),
        LedgerEntry(
            entry_id="LE-002",
            account_code="其他货币资金",  # Account name from mapping rules
            account_path="资产 > 其他货币资金",
            amount=Decimal("500.00"),
            date=datetime(2024, 1, 16),
            description="Test other funds",
            source_entry_id="JE-002",
            rule_applied="R-002",
            quarter=1,
            year=2024,
            ledger_type="DR",
        ),
        LedgerEntry(
            entry_id="LE-003",
            account_code="收入OL",  # Account name from mapping rules
            account_path="收入 > 收入OL",
            amount=Decimal("2000.00"),
            date=datetime(2024, 1, 17),
            description="Test OL income",
            source_entry_id="JE-003",
            rule_applied="R-003",
            quarter=1,
            year=2024,
            ledger_type="CR",
        ),
    ]
    
    output_path = tmp_path / "test_ledger_id_lookup.xlsx"
    generator.generate(ledger_entries, output_path)
    
    # Read back and verify ledger IDs
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Ledger Entries"]
    
    # Verify headers
    assert ws.cell(row=1, column=2).value == "Ledger ID"
    assert ws.cell(row=1, column=3).value == "Account Code"
    
    # Verify data - ledger IDs should be 4-digit numbers
    # Entries may be reordered by hierarchy, so check by account code instead of row number
    ledger_id_map = {}
    for row in range(2, min(ws.max_row + 1, 10)):
        account_code = ws.cell(row=row, column=3).value
        ledger_id = ws.cell(row=row, column=2).value
        if account_code:
            ledger_id_map[account_code] = ledger_id
    
    # Verify each account has the correct ledger ID
    assert ledger_id_map.get("银行存款") == "1100", f"银行存款 should have ledger ID 1100, got {ledger_id_map.get('银行存款')}"
    assert ledger_id_map.get("其他货币资金") == "1200", f"其他货币资金 should have ledger ID 1200, got {ledger_id_map.get('其他货币资金')}"
    assert ledger_id_map.get("收入OL") == "4301", f"收入OL should have ledger ID 4301, got {ledger_id_map.get('收入OL')}"
    
    # Verify all ledger IDs are 4-digit numbers (not account names)
    for account_code, ledger_id in ledger_id_map.items():
        assert isinstance(ledger_id, (str, int)), f"Ledger ID for {account_code} should be string or int, got {type(ledger_id)}"
        ledger_id_str = str(int(ledger_id)) if isinstance(ledger_id, float) else str(ledger_id)
        assert ledger_id_str.isdigit() and len(ledger_id_str) == 4, f"Ledger ID for {account_code} should be 4-digit number, got {ledger_id_str}"


def test_ledger_id_lookup_in_quarterly_report(real_world_hierarchy, tmp_path):
    """Test that ledger ID (4-digit) is correctly looked up and displayed in quarterly report."""
    generator = QuarterlyReportGenerator(real_world_hierarchy)
    
    ledger_entries = [
        LedgerEntry(
            entry_id="LE-001",
            account_code="银行存款",  # Account name from mapping rules
            account_path="资产 > 银行存款",
            amount=Decimal("1000.00"),
            date=datetime(2024, 1, 15),
            description="Test",
            source_entry_id="JE-001",
            rule_applied="R-001",
            quarter=1,
            year=2024,
            ledger_type="CR",
        ),
        LedgerEntry(
            entry_id="LE-002",
            account_code="支出OL讲师",  # Account name from mapping rules
            account_path="支出 > 支出OL讲师",
            amount=Decimal("500.00"),
            date=datetime(2024, 1, 16),
            description="Test",
            source_entry_id="JE-002",
            rule_applied="R-002",
            quarter=1,
            year=2024,
            ledger_type="DR",
        ),
    ]
    
    output_path = tmp_path / "test_quarterly_ledger_id_lookup.xlsx"
    generator.generate(ledger_entries, output_path)
    
    # Read back and verify ledger IDs
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Quarterly Totals"]
    
    # Verify headers
    assert ws.cell(row=3, column=1).value == "Ledger ID"
    assert ws.cell(row=3, column=2).value == "Account Code"
    
    # Find the data rows (skip quarter headers)
    # Look for rows with ledger IDs
    found_ledger_ids = []
    for row in range(4, min(ws.max_row + 1, 20)):
        ledger_id = ws.cell(row=row, column=1).value
        account_code = ws.cell(row=row, column=2).value
        if ledger_id and account_code:
            found_ledger_ids.append((row, ledger_id, account_code))
    
    # Verify we found ledger IDs (4-digit numbers)
    ledger_id_values = [lid for _, lid, _ in found_ledger_ids]
    assert "1100" in ledger_id_values or "5301" in ledger_id_values, f"Expected 1100 or 5301 in ledger IDs, got {ledger_id_values}"






