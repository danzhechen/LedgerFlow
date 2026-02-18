"""
Integration test: run pipeline on journal_entry_2020_2024.xlsx (one sheet) and optionally
validate yearly report numbers against reference Veritas China Account book files.

To run full validation against reference books, set environment variables or pass paths:
  REFERENCE_2020, REFERENCE_2021, REFERENCE_2022, REFERENCE_2023, REFERENCE_2024
or use scripts/validate_yearly_against_reference.py with --reference-YYYY.
"""

from pathlib import Path

import pytest

from veritas_accounting.cli.processor import ProcessingPipeline
from veritas_accounting.config.settings import AppConfig, InputConfig, OutputConfig, ValidationConfig


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


@pytest.mark.integration
def test_journal_2020_2024_sheet_2024_pipeline(tmp_path: Path) -> None:
    """
    Run pipeline on journal_entry_2020_2024.xlsx, sheet "2024".
    Verifies unified report is produced and Account Summary (by Year) has data for 2024.
    """
    root = _project_root()
    journal_file = root / "examples" / "journal_entry_2020_2024.xlsx"
    rules_file = root / "账目分类明细_ledger_rules.xlsx"
    hierarchy_file = root / "账目分类明细.xlsx"

    if not journal_file.exists():
        pytest.skip(f"Journal file not found: {journal_file}")
    if not rules_file.exists():
        pytest.skip(f"Rules file not found: {rules_file}")
    if not hierarchy_file.exists():
        pytest.skip(f"Hierarchy file not found: {hierarchy_file}")

    config = AppConfig(
        input=InputConfig(
            journal_file=str(journal_file),
            rules_file=str(rules_file),
            account_hierarchy_file=str(hierarchy_file),
        ),
        output=OutputConfig(directory=str(tmp_path)),
        validation=ValidationConfig(level="strict", auto_fix_enabled=False),
        sheet_name="2024",
    )
    pipeline = ProcessingPipeline(config)
    result = pipeline.process()

    assert result["success"] is True, result.get("errors", [])
    assert len(result["output_files"]) >= 1

    ledger_path = tmp_path / config.output.ledger_file
    assert ledger_path.exists()

    from openpyxl import load_workbook
    wb = load_workbook(ledger_path)
    assert "Account Summary (by Year)" in wb.sheetnames
    ws = wb["Account Summary (by Year)"]
    # Header row 3, data from row 4; column A = Year
    years_found = set()
    for row in range(4, min(ws.max_row + 1, 200)):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip().upper() != "TOTAL":
            try:
                years_found.add(int(float(val)))
            except (ValueError, TypeError):
                pass
    assert 2024 in years_found, f"Expected year 2024 in Account Summary (by Year). Years found: {years_found}"


@pytest.mark.integration
def test_yearly_validation_against_reference_if_provided(tmp_path: Path) -> None:
    """
    If REFERENCE_2024 is set and pipeline output exists for 2024, compare yearly numbers.
    Run pipeline first for sheet 2024, then compare with reference file.
    """
    import os
    import sys
    # Allow importing yearly_validation_helpers from scripts/
    scripts_dir = _project_root() / "scripts"
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    from yearly_validation_helpers import (
        compare_yearly,
        get_yearly_numbers_from_pipeline_output,
        get_yearly_numbers_from_reference,
    )

    ref_path = os.environ.get("REFERENCE_2024")
    if not ref_path or not Path(ref_path).exists():
        pytest.skip("REFERENCE_2024 not set or file not found (optional validation)")

    root = _project_root()
    journal_file = root / "examples" / "journal_entry_2020_2024.xlsx"
    rules_file = root / "账目分类明细_ledger_rules.xlsx"
    hierarchy_file = root / "账目分类明细.xlsx"
    if not all(p.exists() for p in (journal_file, rules_file, hierarchy_file)):
        pytest.skip("Missing journal/rules/hierarchy files")

    config = AppConfig(
        input=InputConfig(
            journal_file=str(journal_file),
            rules_file=str(rules_file),
            account_hierarchy_file=str(hierarchy_file),
        ),
        output=OutputConfig(directory=str(tmp_path)),
        validation=ValidationConfig(level="strict", auto_fix_enabled=False),
        sheet_name="2024",
    )
    pipeline = ProcessingPipeline(config)
    result = pipeline.process()
    assert result["success"] is True

    our_report = tmp_path / config.output.ledger_file
    ours = get_yearly_numbers_from_pipeline_output(our_report, 2024)
    reference = get_yearly_numbers_from_reference(Path(ref_path))
    assert reference, "Reference file has no readable data"

    diffs = compare_yearly(ours, reference, tolerance=0.01)
    assert not diffs, f"Yearly numbers differ from reference: {diffs[:20]}"
