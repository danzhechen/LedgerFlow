"""Integration test for examples/test_journal_new.xlsx pipeline."""

from pathlib import Path

from veritas_accounting.cli.processor import ProcessingPipeline
from veritas_accounting.config.settings import AppConfig, InputConfig, OutputConfig


def test_test_journal_new_pipeline(tmp_path):
    """
    Run the full processing pipeline on examples/test_journal_new.xlsx.

    Uses: journal_file=examples/test_journal_new.xlsx, rules=账目分类明细_ledger_rules.xlsx,
    account_hierarchy=账目分类明细.xlsx.
    """

    project_root = Path(__file__).resolve().parents[2]

    journal_file = project_root / "examples" / "test_journal_new.xlsx"
    rules_file = project_root / "账目分类明细_ledger_rules.xlsx"
    hierarchy_file = project_root / "账目分类明细.xlsx"
    output_dir = tmp_path  # isolated output per test run

    # Sanity checks so the test fails clearly if files are missing
    assert journal_file.exists(), f"Missing journal file: {journal_file}"
    assert rules_file.exists(), f"Missing rules file: {rules_file}"
    assert hierarchy_file.exists(), f"Missing hierarchy file: {hierarchy_file}"

    config = AppConfig(
        input=InputConfig(
            journal_file=str(journal_file),
            rules_file=str(rules_file),
            account_hierarchy_file=str(hierarchy_file),
        ),
        output=OutputConfig(
            directory=str(output_dir),
        ),
    )

    pipeline = ProcessingPipeline(config)
    result = pipeline.process()

    # Basic assertions (unified report: single file with multiple sheets)
    assert result["success"] is True
    assert len(result["output_files"]) >= 1

    # Single unified report file
    ledger_path = output_dir / config.output.ledger_file
    assert ledger_path.exists(), f"Expected output not found: {config.output.ledger_file}"

    from openpyxl import load_workbook

    wb = load_workbook(ledger_path)
    # Unified report sheets: Ledger Entries, Account Summary (by Year), Account Summary (by Quarter), Quarterly Report, Audit & Review
    assert "Ledger Entries" in wb.sheetnames, f"Ledger Entries sheet missing. Sheets: {wb.sheetnames}"
    ws = wb["Ledger Entries"]
    
    # Verify header has "Ledger ID" column
    assert ws.cell(row=1, column=2).value == "Ledger ID", f"Ledger ID column missing in header. Found: {[ws.cell(row=1, col=i).value for i in range(1, 6)]}"
    
    # Debug: Print first few rows to see what's actually there
    print(f"\nDEBUG: Ledger Entries sheet - Max row: {ws.max_row}")
    print(f"Headers: {[ws.cell(row=1, column=i).value for i in range(1, 6)]}")
    for row in range(2, min(ws.max_row + 1, 5)):
        row_data = [ws.cell(row=row, column=i).value for i in range(1, 6)]
        print(f"Row {row}: {row_data}")
    
    # Check that at least one row has a ledger ID (4-digit number)
    # Ledger ID should be a 4-digit number, not account names
    has_ledger_id = False
    ledger_ids_found = []
    for row in range(2, min(ws.max_row + 1, 20)):  # Check first 18 data rows
        ledger_id = ws.cell(row=row, column=2).value
        if ledger_id:
            ledger_ids_found.append((row, ledger_id, type(ledger_id).__name__))
            # Ledger ID should be a 4-digit number (string representation)
            if isinstance(ledger_id, (str, int, float)):
                ledger_id_str = str(int(ledger_id)) if isinstance(ledger_id, float) else str(ledger_id)
                if ledger_id_str.isdigit() and len(ledger_id_str) == 4:
                    has_ledger_id = True
                    print(f"✓ Found valid ledger ID at row {row}: {ledger_id_str}")
                    break
    
    if not has_ledger_id:
        print(f"DEBUG: Ledger IDs found (first 10): {ledger_ids_found[:10]}")
        print(f"DEBUG: Account codes (col 3, first 5 rows): {[ws.cell(row=i, column=3).value for i in range(2, min(ws.max_row+1, 7))]}")
    
    assert has_ledger_id, f"Ledger ID (4-digit number) not found in output. Found values: {ledger_ids_found[:10]}"
    
    # Check quarterly report (sheet in same workbook)
    assert "Quarterly Report" in wb.sheetnames, f"Quarterly Report sheet missing. Sheets: {wb.sheetnames}"
    ws_q = wb["Quarterly Report"]
    
    # Verify header has "Ledger ID" column
    assert ws_q.cell(row=3, column=1).value == "Ledger ID", f"Ledger ID column missing in quarterly report header. Found: {[ws_q.cell(row=3, col=i).value for i in range(1, 6)]}"
    
    # Debug quarterly report
    print(f"\nDEBUG: Quarterly Totals sheet - Max row: {ws_q.max_row}")
    print(f"Headers (row 3): {[ws_q.cell(row=3, column=i).value for i in range(1, 6)]}")
    for row in range(4, min(ws_q.max_row + 1, 8)):
        row_data = [ws_q.cell(row=row, column=i).value for i in range(1, 6)]
        print(f"Row {row}: {row_data}")
    
    # Check that at least one row has a ledger ID
    has_ledger_id_q = False
    ledger_ids_found_q = []
    for row in range(4, min(ws_q.max_row + 1, 20)):  # Check first data rows
        ledger_id = ws_q.cell(row=row, column=1).value
        if ledger_id:
            ledger_ids_found_q.append((row, ledger_id, type(ledger_id).__name__))
            # Ledger ID should be a 4-digit number
            if isinstance(ledger_id, (str, int, float)):
                ledger_id_str = str(int(ledger_id)) if isinstance(ledger_id, float) else str(ledger_id)
                if ledger_id_str.isdigit() and len(ledger_id_str) == 4:
                    has_ledger_id_q = True
                    print(f"✓ Found valid ledger ID in quarterly report at row {row}: {ledger_id_str}")
                    break
    
    if not has_ledger_id_q:
        print(f"DEBUG: Ledger IDs found in quarterly (first 10): {ledger_ids_found_q[:10]}")
    
    assert has_ledger_id_q, f"Ledger ID (4-digit number) not found in quarterly report. Found values: {ledger_ids_found_q[:10]}"

    # Audit & Review sheet in same workbook
    assert "Audit & Review" in wb.sheetnames, "Audit & Review sheet missing"
    ws_audit = wb["Audit & Review"]
    assert ws_audit.max_row >= 2, "Audit & Review should have content"







