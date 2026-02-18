"""Audit trail export for veritas-accounting."""

import csv
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from veritas_accounting.audit.trail import AuditTrail
from veritas_accounting.utils.exceptions import ExcelIOError


class AuditTrailExporter:
    """
    Exports audit trail data to multiple formats.

    Supports Excel, CSV, and JSON export formats for external analysis
    and compliance record keeping.
    """

    def __init__(self, audit_trail: AuditTrail) -> None:
        """
        Initialize AuditTrailExporter.

        Args:
            audit_trail: AuditTrail to export
        """
        self.audit_trail = audit_trail

    def export_to_excel(self, output_path: Path | str) -> None:
        """
        Export audit trail to Excel format.

        Args:
            output_path: Path to output Excel file

        Raises:
            ExcelIOError: If export fails
        """
        output_path = Path(output_path)

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create sheets
            self._create_metadata_sheet(wb)
            self._create_transformations_sheet(wb)
            self._create_rules_sheet(wb)
            self._create_entries_sheet(wb)
            self._create_relationships_sheet(wb)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

        except Exception as e:
            raise ExcelIOError(
                f"Failed to export audit trail to Excel: {output_path}. Error: {e}"
            ) from e

    def export_to_csv(self, output_path: Path | str) -> None:
        """
        Export audit trail to CSV format.

        Args:
            output_path: Path to output CSV file (or directory for multiple files)

        Raises:
            IOError: If export fails
        """
        output_path = Path(output_path)

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Export transformations as main CSV
            transformations_data = []
            for record in self.audit_trail.records:
                for ledger_entry in record.generated_entries:
                    transformations_data.append({
                        "journal_entry_id": record.entry_id,
                        "transformation_timestamp": record.timestamp.isoformat(),
                        "rule_id": ledger_entry.rule_applied,
                        "ledger_entry_id": ledger_entry.entry_id,
                        "account_code": ledger_entry.account_code,
                        "amount": float(ledger_entry.amount),
                        "date": ledger_entry.date.isoformat(),
                        "no_match": record.no_match,
                        "user": record.user or "",
                        "system_version": record.system_version or "",
                        "rule_version": record.rule_version or "",
                    })

            if transformations_data:
                df = pd.DataFrame(transformations_data)
                df.to_csv(output_path, index=False, encoding="utf-8-sig")

        except Exception as e:
            raise IOError(
                f"Failed to export audit trail to CSV: {output_path}. Error: {e}"
            ) from e

    def export_to_json(self, output_path: Path | str) -> None:
        """
        Export audit trail to JSON format.

        Args:
            output_path: Path to output JSON file

        Raises:
            IOError: If export fails
        """
        output_path = Path(output_path)

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Export full audit trail as JSON
            audit_data = self.audit_trail.to_dict()

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(audit_data, f, indent=2, ensure_ascii=False, default=str)

        except Exception as e:
            raise IOError(
                f"Failed to export audit trail to JSON: {output_path}. Error: {e}"
            ) from e

    def _create_metadata_sheet(self, wb: Workbook) -> None:
        """Create metadata sheet with audit trail information."""
        ws = wb.create_sheet("Metadata", 0)

        # Title
        ws["A1"] = "Audit Trail Metadata"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        row = 3
        metadata = self.audit_trail.to_dict()["metadata"]
        summary = self.audit_trail.get_summary()

        # Metadata
        for key, value in metadata.items():
            if value:
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"B{row}"] = str(value)
                row += 1

        # Summary stats
        row += 1
        ws[f"A{row}"] = "Summary Statistics"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        for key, value in summary.items():
            if key not in ["start_time", "end_time", "duration_seconds"]:
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"B{row}"] = str(value)
                row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_transformations_sheet(self, wb: Workbook) -> None:
        """Create transformations sheet with all transformation records."""
        ws = wb.create_sheet("Transformations")

        # Headers
        headers = [
            "Entry ID",
            "Timestamp",
            "User",
            "System Version",
            "Rule Version",
            "Source Description",
            "Source Amount",
            "Source Date",
            "Applied Rules",
            "Generated Entries Count",
            "No Match",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="2C3E50",
                end_color="2C3E50",
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Transformation data
        for row_idx, record in enumerate(self.audit_trail.records, start=2):
            ws.cell(row=row_idx, column=1, value=record.entry_id)
            ws.cell(row=row_idx, column=2, value=record.timestamp.isoformat())
            ws.cell(row=row_idx, column=3, value=record.user or "")
            ws.cell(row=row_idx, column=4, value=record.system_version or "")
            ws.cell(row=row_idx, column=5, value=record.rule_version or "")
            ws.cell(row=row_idx, column=6, value=record.source_entry.description)
            ws.cell(row=row_idx, column=7, value=float(record.source_entry.amount))
            ws.cell(row=row_idx, column=8, value=record.source_entry.date.isoformat())
            ws.cell(
                row=row_idx,
                column=9,
                value=", ".join(rule.rule_id for rule in record.applied_rules),
            )
            ws.cell(row=row_idx, column=10, value=len(record.generated_entries))
            ws.cell(row=row_idx, column=11, value="Yes" if record.no_match else "No")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_rules_sheet(self, wb: Workbook) -> None:
        """Create rules sheet with all applied rules."""
        ws = wb.create_sheet("Rules")

        # Collect unique rules
        unique_rules: dict[str, dict[str, Any]] = {}
        for record in self.audit_trail.records:
            for rule in record.applied_rules:
                if rule.rule_id not in unique_rules:
                    unique_rules[rule.rule_id] = {
                        "rule_id": rule.rule_id,
                        "condition": rule.condition,
                        "account_code": rule.account_code,
                        "description": rule.description,
                        "priority": rule.priority,
                        "generates_multiple": rule.generates_multiple,
                        "usage_count": 0,
                    }
                unique_rules[rule.rule_id]["usage_count"] += 1

        # Headers
        headers = [
            "Rule ID",
            "Condition",
            "Account Code",
            "Description",
            "Priority",
            "Generates Multiple",
            "Usage Count",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="2C3E50",
                end_color="2C3E50",
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Rule data
        for row_idx, rule_data in enumerate(unique_rules.values(), start=2):
            ws.cell(row=row_idx, column=1, value=rule_data["rule_id"])
            ws.cell(row=row_idx, column=2, value=rule_data["condition"])
            ws.cell(row=row_idx, column=3, value=rule_data["account_code"])
            ws.cell(row=row_idx, column=4, value=rule_data["description"] or "")
            ws.cell(row=row_idx, column=5, value=rule_data["priority"])
            ws.cell(row=row_idx, column=6, value="Yes" if rule_data["generates_multiple"] else "No")
            ws.cell(row=row_idx, column=7, value=rule_data["usage_count"])

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_entries_sheet(self, wb: Workbook) -> None:
        """Create entries sheet with all journal and ledger entries."""
        ws = wb.create_sheet("Entries")

        # Headers
        headers = [
            "Entry Type",
            "Entry ID",
            "Source Entry ID",
            "Account Code",
            "Amount",
            "Date",
            "Description",
            "Rule Applied",
            "Quarter",
            "Year",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="2C3E50",
                end_color="2C3E50",
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Entry data (journal entries first, then ledger entries)
        row_idx = 2
        for record in self.audit_trail.records:
            # Journal entry
            ws.cell(row=row_idx, column=1, value="Journal")
            ws.cell(row=row_idx, column=2, value=record.source_entry.entry_id)
            ws.cell(row=row_idx, column=3, value="")
            ws.cell(row=row_idx, column=4, value=record.source_entry.old_type)
            ws.cell(row=row_idx, column=5, value=float(record.source_entry.amount))
            ws.cell(row=row_idx, column=6, value=record.source_entry.date.isoformat())
            ws.cell(row=row_idx, column=7, value=record.source_entry.description)
            ws.cell(row=row_idx, column=8, value="")
            ws.cell(row=row_idx, column=9, value=record.source_entry.quarter or "")
            ws.cell(row=row_idx, column=10, value=record.source_entry.year)
            row_idx += 1

            # Ledger entries
            for ledger_entry in record.generated_entries:
                ws.cell(row=row_idx, column=1, value="Ledger")
                ws.cell(row=row_idx, column=2, value=ledger_entry.entry_id)
                ws.cell(row=row_idx, column=3, value=ledger_entry.source_entry_id)
                ws.cell(row=row_idx, column=4, value=ledger_entry.account_code)
                ws.cell(row=row_idx, column=5, value=float(ledger_entry.amount))
                ws.cell(row=row_idx, column=6, value=ledger_entry.date.isoformat())
                ws.cell(row=row_idx, column=7, value=ledger_entry.description)
                ws.cell(row=row_idx, column=8, value=ledger_entry.rule_applied)
                ws.cell(row=row_idx, column=9, value=ledger_entry.quarter)
                ws.cell(row=row_idx, column=10, value=ledger_entry.year)
                row_idx += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_relationships_sheet(self, wb: Workbook) -> None:
        """Create relationships sheet showing links between entries and rules."""
        ws = wb.create_sheet("Relationships")

        # Headers
        headers = [
            "Journal Entry ID",
            "Rule ID",
            "Ledger Entry ID",
            "Account Code",
            "Amount",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="2C3E50",
                end_color="2C3E50",
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Relationship data
        row_idx = 2
        for record in self.audit_trail.records:
            for ledger_entry in record.generated_entries:
                ws.cell(row=row_idx, column=1, value=record.entry_id)
                ws.cell(row=row_idx, column=2, value=ledger_entry.rule_applied)
                ws.cell(row=row_idx, column=3, value=ledger_entry.entry_id)
                ws.cell(row=row_idx, column=4, value=ledger_entry.account_code)
                ws.cell(row=row_idx, column=5, value=float(ledger_entry.amount))
                row_idx += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions








