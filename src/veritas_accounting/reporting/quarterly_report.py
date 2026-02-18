"""Quarterly aggregation report generation for veritas-accounting."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from veritas_accounting.models.account import AccountHierarchy
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.transformation.aggregator import (
    QuarterlyAggregator,
    QuarterlyAggregation,
    HierarchicalTotals,
)
from veritas_accounting.utils.exceptions import ExcelIOError


class QuarterlyReportGenerator:
    """
    Generates quarterly aggregation reports in Excel format.

    Creates professional Excel files with quarterly totals, hierarchical summaries,
    and statistics ready for financial reporting.
    """

    # Color scheme
    COLOR_HEADER = "2C3E50"  # Dark blue-gray
    COLOR_TOTAL = "3498DB"  # Blue
    COLOR_QUARTER1 = "E74C3C"  # Red
    COLOR_QUARTER2 = "F39C12"  # Orange
    COLOR_QUARTER3 = "2ECC71"  # Green
    COLOR_QUARTER4 = "9B59B6"  # Purple

    def __init__(
        self, account_hierarchy: AccountHierarchy | None = None
    ) -> None:
        """
        Initialize QuarterlyReportGenerator.

        Args:
            account_hierarchy: AccountHierarchy object (optional, for hierarchical organization)
        """
        self.account_hierarchy = account_hierarchy
        self.aggregator = QuarterlyAggregator(account_hierarchy)

    def generate(
        self,
        ledger_entries: list[LedgerEntry],
        output_path: Path | str,
        include_charts: bool = True,
    ) -> None:
        """
        Generate quarterly aggregation report Excel file.

        Args:
            ledger_entries: List of LedgerEntry objects
            output_path: Path to output Excel file
            include_charts: Whether to include charts (default: True)

        Raises:
            ExcelIOError: If generation fails
        """
        output_path = Path(output_path)

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Generate aggregations
            aggregations = self.aggregator.aggregate(ledger_entries)
            level_totals = self.aggregator.aggregate_by_level(ledger_entries)

            # Create sheets
            self._create_quarterly_totals_sheet(wb, aggregations)
            self._create_hierarchy_summary_sheet(wb, aggregations)
            self._create_statistics_sheet(wb, ledger_entries, aggregations)

            # Add charts if requested
            if include_charts:
                self._add_charts(wb, aggregations)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

        except Exception as e:
            raise ExcelIOError(
                f"Failed to generate quarterly report: {output_path}. Error: {e}"
            ) from e

    def add_quarterly_totals_sheet(
        self,
        wb: Workbook,
        ledger_entries: list[LedgerEntry],
        index: int,
    ) -> None:
        """
        Add the Quarterly Report sheet to an existing workbook.
        Used when building a unified report in a single file.
        """
        aggregations = self.aggregator.aggregate(ledger_entries)
        ws = wb.create_sheet("Quarterly Report", index)
        self._fill_quarterly_totals_sheet(ws, aggregations)

    def _create_quarterly_totals_sheet(
        self, wb: Workbook, aggregations: list[QuarterlyAggregation]
    ) -> None:
        """Create quarterly totals sheet."""
        ws = wb.create_sheet("Quarterly Totals", 0)
        self._fill_quarterly_totals_sheet(ws, aggregations)

    def _fill_quarterly_totals_sheet(
        self, ws: Any, aggregations: list[QuarterlyAggregation]
    ) -> None:
        """Fill worksheet with quarterly totals (title + data)."""
        # Title
        ws["A1"] = "Quarterly Aggregation Totals"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        # Headers
        headers = [
            "Ledger ID",  # 4-digit number from ledger_new section (e.g., 1100, 2209)
            "Account Code",  # Account code from mapping rules
            "Account Name",
            "Quarter",
            "Year",
            "CR Amount",
            "DR Amount",
            "Net Amount",
            "Entry Count",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Group by quarter and year for comparison
        by_quarter: dict[tuple[int, int], list[QuarterlyAggregation]] = {}
        for agg in aggregations:
            key = (agg.quarter, agg.year)
            if key not in by_quarter:
                by_quarter[key] = []
            by_quarter[key].append(agg)

        # Write data
        row = 4
        for (quarter, year), aggs in sorted(by_quarter.items()):
            # Quarter header
            quarter_label = f"Q{quarter} {year}"
            ws.cell(row=row, column=1, value=quarter_label)
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:I{row}")
            row += 1

            # Quarter data - sort by account code (simplified, no hierarchy levels)
            for agg in sorted(aggs, key=lambda a: a.account_code):
                # Get account name and ledger ID from hierarchy if available
                account_name = agg.account_path
                ledger_id = agg.account_code  # Default to account_code
                if self.account_hierarchy:
                    account = self.account_hierarchy.get_account(agg.account_code)
                    # If not found by code, try looking up by name (mapping rules might use names)
                    if not account:
                        account = self.account_hierarchy.get_account_by_name(agg.account_code)
                    if account:
                        account_name = account.name
                        ledger_id = account.code  # This is the 4-digit ledger ID from ledger_new
                
                ws.cell(row=row, column=1, value=ledger_id)  # Ledger ID (4-digit number from ledger_new)
                ws.cell(row=row, column=2, value=agg.account_code)  # Account Code
                ws.cell(row=row, column=3, value=account_name)
                ws.cell(row=row, column=4, value=agg.quarter)
                ws.cell(row=row, column=5, value=agg.year)
                ws.cell(row=row, column=6, value=float(agg.cr_amount))
                ws.cell(row=row, column=7, value=float(agg.dr_amount))
                ws.cell(row=row, column=8, value=float(agg.total_amount))
                ws.cell(row=row, column=9, value=agg.entry_count)

                # Format amount cells
                for col in [6, 7, 8]:  # CR, DR, Net columns
                    amount_cell = ws.cell(row=row, column=col)
                    amount_cell.number_format = "#,##0.00"

                # Color code by quarter
                quarter_color = self._get_quarter_color(agg.quarter)
                if quarter_color:
                    for col in range(1, 10):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(
                            start_color=quarter_color,
                            end_color=quarter_color,
                            fill_type="solid",
                        )

                row += 1

            # Quarter totals
            quarter_cr_total = sum(float(a.cr_amount) for a in aggs)
            quarter_dr_total = sum(float(a.dr_amount) for a in aggs)
            quarter_net_total = sum(float(a.total_amount) for a in aggs)
            quarter_count = sum(a.entry_count for a in aggs)
            # Net for TOTAL row: show DR - CR (balanced ledger => 0).
            quarter_net_display = float(quarter_dr_total) - float(quarter_cr_total)

            ws.cell(row=row, column=1, value=f"{quarter_label} TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.merge_cells(f"A{row}:E{row}")
            ws.cell(row=row, column=6, value=float(quarter_cr_total))
            ws.cell(row=row, column=7, value=float(quarter_dr_total))
            ws.cell(row=row, column=8, value=quarter_net_display)
            ws.cell(row=row, column=9, value=quarter_count)

            # Format total row
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.COLOR_TOTAL,
                    end_color=self.COLOR_TOTAL,
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color="FFFFFF")
                if col in [6, 7, 8]:  # CR, DR, Net columns
                    cell.number_format = "#,##0.00"

            row += 1

            # 收入−支出 = Σ(CR−DR) for 收入(4xxx) + 支出(5xxx). Should equal 小结 期末总净资产−期初总净资产.
            # If sign is opposite to 小结, check mapping rules: 收入科目应为 CR、支出科目应为 DR（正数交易）.
            quarter_result_ie_raw = self._quarter_result_by_code_prefix(aggs, ("4", "5"))
            quarter_result_ie = float(quarter_result_ie_raw)
            quarter_result_equity = self._quarter_result_by_code_prefix(aggs, ("3",))
            ws.cell(row=row, column=1, value=f"{quarter_label} Result (收入−支出)")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.merge_cells(f"A{row}:E{row}")
            ws.cell(row=row, column=8, value=quarter_result_ie)
            for col in [6, 7, 9]:
                ws.cell(row=row, column=col, value=None)
            for col in [1, 8]:
                ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=8).number_format = "#,##0.00"
            row += 1
            ws.cell(row=row, column=1, value=f"{quarter_label} Result (支出−收入)")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.merge_cells(f"A{row}:E{row}")
            ws.cell(row=row, column=8, value=float(-quarter_result_ie))
            for col in [6, 7, 9]:
                ws.cell(row=row, column=col, value=None)
            for col in [1, 8]:
                ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=8).number_format = "#,##0.00"
            row += 1
            ws.cell(row=row, column=1, value=f"{quarter_label} Result (权益)")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.merge_cells(f"A{row}:E{row}")
            ws.cell(row=row, column=8, value=float(quarter_result_equity))
            for col in [6, 7, 9]:
                ws.cell(row=row, column=col, value=None)
            for col in [1, 8]:
                ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=8).number_format = "#,##0.00"
            row += 2  # Space between quarters

        # Auto-adjust column widths
        column_widths = {
            "A": 15,  # Ledger ID
            "B": 15,  # Account Code
            "C": 40,  # Account Name
            "D": 8,   # Quarter
            "E": 8,   # Year
            "F": 18,  # CR Amount
            "G": 18,  # DR Amount
            "H": 18,  # Net Amount
            "I": 12,  # Entry Count
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top rows
        ws.freeze_panes = "A4"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_hierarchy_summary_sheet(
        self, wb: Workbook, aggregations: list[QuarterlyAggregation]
    ) -> None:
        """Create simplified account summary sheet (no heavy hierarchy)."""
        ws = wb.create_sheet("Account Summary")

        # Title
        ws["A1"] = "Account Summary (All Accounts)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        # Headers
        headers = [
            "Ledger ID",  # 4-digit number from ledger_new section (e.g., 1100, 2209)
            "Account Code",  # Account code from mapping rules
            "Account Name",
            "CR Amount",
            "DR Amount",
            "Net Amount",
            "Entry Count",
        ]
        row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        row += 1

        # Get all aggregations and group by account_code (sum across all quarters)
        from collections import defaultdict
        from decimal import Decimal
        
        account_totals: dict[str, dict] = defaultdict(lambda: {
            "cr_amount": Decimal("0"),
            "dr_amount": Decimal("0"),
            "total_amount": Decimal("0"),
            "entry_count": 0,
            "account_path": "",
        })
        
        # Sum aggregations by account_code across all quarters
        for agg in aggregations:
            code = agg.account_code
            account_totals[code]["cr_amount"] += agg.cr_amount
            account_totals[code]["dr_amount"] += agg.dr_amount
            account_totals[code]["total_amount"] += agg.total_amount
            account_totals[code]["entry_count"] += agg.entry_count
            if not account_totals[code]["account_path"]:
                account_totals[code]["account_path"] = agg.account_path
        
        # If no data, show message
        if not account_totals:
            ws.cell(row=row, column=1, value="No data available.")
            return

        # Write account summary data
        for account_code in sorted(account_totals.keys()):
            totals = account_totals[account_code]
            
            # Get account name and ledger ID from hierarchy if available
            account_name = totals["account_path"]
            ledger_id = account_code  # Default to account_code
            if self.account_hierarchy:
                account = self.account_hierarchy.get_account(account_code)
                # If not found by code, try looking up by name (mapping rules might use names)
                if not account:
                    account = self.account_hierarchy.get_account_by_name(account_code)
                if account:
                    account_name = account.name
                    ledger_id = account.code  # This is the 4-digit ledger ID from ledger_new
            
            ws.cell(row=row, column=1, value=ledger_id)  # Ledger ID (4-digit number from ledger_new)
            ws.cell(row=row, column=2, value=account_code)  # Account Code
            ws.cell(row=row, column=3, value=account_name)
            ws.cell(row=row, column=4, value=float(totals["cr_amount"]))
            ws.cell(row=row, column=5, value=float(totals["dr_amount"]))
            ws.cell(row=row, column=6, value=float(totals["total_amount"]))
            ws.cell(row=row, column=7, value=totals["entry_count"])

            # Format amount cells
            for col in [4, 5, 6]:  # CR, DR, Net columns
                amount_cell = ws.cell(row=row, column=col)
                amount_cell.number_format = "#,##0.00"

            row += 1

        # Grand total
        grand_cr = sum(float(t["cr_amount"]) for t in account_totals.values())
        grand_dr = sum(float(t["dr_amount"]) for t in account_totals.values())
        grand_net = sum(float(t["total_amount"]) for t in account_totals.values())
        grand_count = sum(t["entry_count"] for t in account_totals.values())

        ws.cell(row=row, column=1, value="GRAND TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=4, value=grand_cr)
        ws.cell(row=row, column=5, value=grand_dr)
        ws.cell(row=row, column=6, value=grand_net)
        ws.cell(row=row, column=7, value=grand_count)

        # Format total row
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_TOTAL,
                end_color=self.COLOR_TOTAL,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")
            if col in [4, 5, 6]:  # CR, DR, Net columns
                cell.number_format = "#,##0.00"

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 15  # Ledger ID
        ws.column_dimensions["B"].width = 15  # Account Code
        ws.column_dimensions["C"].width = 40  # Account Name
        ws.column_dimensions["D"].width = 18  # CR Amount
        ws.column_dimensions["E"].width = 18  # DR Amount
        ws.column_dimensions["F"].width = 18  # Net Amount
        ws.column_dimensions["G"].width = 15  # Entry Count

    def _create_statistics_sheet(
        self,
        wb: Workbook,
        ledger_entries: list[LedgerEntry],
        aggregations: list[QuarterlyAggregation],
    ) -> None:
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")

        # Title
        ws["A1"] = "Summary Statistics"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        # Calculate statistics
        total_entries = len(ledger_entries)
        total_amount = sum(float(e.amount) for e in ledger_entries)
        unique_accounts = len(set(e.account_code for e in ledger_entries))
        unique_quarters = len(
            set((e.quarter, e.year) for e in ledger_entries)
        )

        # Quarters breakdown
        quarters = {}
        for entry in ledger_entries:
            key = (entry.quarter, entry.year)
            if key not in quarters:
                quarters[key] = {"count": 0, "amount": 0.0}
            quarters[key]["count"] += 1
            quarters[key]["amount"] += float(entry.amount)

        # Write statistics
        row = 3
        stats = [
            ("Total Entries", total_entries),
            ("Total Amount", f"{total_amount:,.2f}"),
            ("Unique Accounts", unique_accounts),
            ("Unique Quarters", unique_quarters),
        ]

        for label, value in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Quarters breakdown
        row += 1
        ws.cell(row=row, column=1, value="Quarter Breakdown")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        headers = ["Quarter", "Entry Count", "Total Amount"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
        row += 1

        for (quarter, year), data in sorted(quarters.items()):
            ws.cell(row=row, column=1, value=f"Q{quarter} {year}")
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=data["amount"])
            ws.cell(row=row, column=3).number_format = "#,##0.00"
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

    def _add_charts(self, wb: Workbook, aggregations: list[QuarterlyAggregation]) -> None:
        """Add charts to the quarterly totals sheet."""
        ws = wb["Quarterly Totals"]

        # Create a simple bar chart for quarterly totals
        # This is a placeholder - actual chart implementation would require
        # more complex data organization
        # Charts can be added using openpyxl.chart module

    def _quarter_result_by_code_prefix(
        self,
        aggs: list[QuarterlyAggregation],
        code_prefixes: tuple[str, ...],
    ) -> float:
        """
        Sum of Net (CR - DR) for accounts whose hierarchy code starts with one of
        code_prefixes. E.g. ("4", "5") for 收入+支出, ("3",) for 权益.
        Matches "end of quarter minus beginning of quarter" for that category.
        """
        from decimal import Decimal

        total = Decimal("0")
        for agg in aggs:
            if not self.account_hierarchy:
                continue
            acc = self.account_hierarchy.get_account(agg.account_code)
            if not acc:
                acc = self.account_hierarchy.get_account_by_name(agg.account_code)
            if acc and acc.code.strip():
                first = acc.code.strip()[0]
                if first in code_prefixes:
                    total += agg.total_amount
        return float(total)

    def _get_quarter_color(self, quarter: int) -> str | None:
        """
        Get color for quarter.

        Args:
            quarter: Quarter number (1-4)

        Returns:
            Color hex code or None
        """
        colors = {
            1: self.COLOR_QUARTER1,
            2: self.COLOR_QUARTER2,
            3: self.COLOR_QUARTER3,
            4: self.COLOR_QUARTER4,
        }
        return colors.get(quarter)






