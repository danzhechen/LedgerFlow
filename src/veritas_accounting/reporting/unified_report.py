"""Unified report generation: single workbook with Ledger, Account Summary, Quarterly Report, and Audit & Review."""

from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from veritas_accounting.audit.trail import AuditTrail
from veritas_accounting.models.account import AccountHierarchy
from veritas_accounting.models.journal import JournalEntry
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.reporting.ledger_output import LedgerOutputGenerator
from veritas_accounting.reporting.quarterly_report import QuarterlyReportGenerator
from veritas_accounting.transformation.aggregator import (
    QuarterlyAggregator,
    QuarterlyAggregation,
)
from veritas_accounting.utils.exceptions import ExcelIOError
from veritas_accounting.validation.error_detector import ErrorDetector


class UnifiedReportGenerator:
    """
    Generates a single Excel workbook containing:
    1. Ledger Entries (detail for eyeball/debug)
    2. Account Summary (by Year) - CR, DR, Net per account per year
    3. Account Summary (by Quarter) - CR, DR, Net per account per quarter per year
    4. Quarterly Report - quarterly breakdown
    5. Audit & Review - summary stats, unmatched entries, and transformation overview
    """

    COLOR_HEADER = "2C3E50"
    COLOR_TOTAL = "3498DB"

    def __init__(self, account_hierarchy: AccountHierarchy | None = None) -> None:
        self.account_hierarchy = account_hierarchy
        self.aggregator = QuarterlyAggregator(account_hierarchy)
        self.ledger_generator = LedgerOutputGenerator(account_hierarchy)
        self.quarterly_generator = QuarterlyReportGenerator(account_hierarchy)

    def generate(
        self,
        ledger_entries: list[LedgerEntry],
        journal_entries: list[JournalEntry],
        audit_trail: AuditTrail,
        error_detector: Optional[ErrorDetector] = None,
        output_path: Path | str = "report.xlsx",
    ) -> None:
        """Generate the unified report workbook and save to output_path."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        # 1. Ledger Entries
        self.ledger_generator.add_ledger_entries_sheet(wb, ledger_entries, index=0)

        # 2. Account Summary by Year
        aggregations = self.aggregator.aggregate(ledger_entries)
        self._add_account_summary_by_year_sheet(wb, aggregations, index=1)

        # 3. Account Summary by Quarter
        self._add_account_summary_by_quarter_sheet(wb, aggregations, index=2)

        # 4. Quarterly Report
        self.quarterly_generator.add_quarterly_totals_sheet(wb, ledger_entries, index=3)

        # 5. Audit & Review
        self._add_audit_review_sheet(
            wb, audit_trail, error_detector, journal_entries, ledger_entries, index=4
        )

        try:
            wb.save(output_path)
        except Exception as e:
            raise ExcelIOError(
                f"Failed to save unified report: {output_path}. Error: {e}"
            ) from e

    def _add_account_summary_by_year_sheet(
        self,
        wb: Workbook,
        aggregations: list[QuarterlyAggregation],
        index: int,
    ) -> None:
        """One row per account per year: Year, Ledger ID, Account Code, Account Name, CR, DR, Net, Entry Count."""
        ws = wb.create_sheet("Account Summary (by Year)", index)
        ws["A1"] = "Account Summary by Year"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:H1")

        headers = [
            "Year",
            "Ledger ID",
            "Account Code",
            "Account Name",
            "CR Amount",
            "DR Amount",
            "Net Amount",
            "Entry Count",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
        row = 4

        # Group by (year, account_code)
        by_year_account: dict[tuple[int, str], dict[str, Any]] = defaultdict(
            lambda: {
                "cr_amount": Decimal("0"),
                "dr_amount": Decimal("0"),
                "total_amount": Decimal("0"),
                "entry_count": 0,
                "account_path": "",
            }
        )
        for agg in aggregations:
            key = (agg.year, agg.account_code)
            by_year_account[key]["cr_amount"] += agg.cr_amount
            by_year_account[key]["dr_amount"] += agg.dr_amount
            by_year_account[key]["total_amount"] += agg.total_amount
            by_year_account[key]["entry_count"] += agg.entry_count
            if not by_year_account[key]["account_path"]:
                by_year_account[key]["account_path"] = agg.account_path

        for (year, account_code), tot in sorted(by_year_account.items()):
            ledger_id = account_code
            account_name = tot["account_path"]
            if self.account_hierarchy:
                acc = self.account_hierarchy.get_account(account_code)
                if not acc:
                    acc = self.account_hierarchy.get_account_by_name(account_code)
                if acc:
                    account_name = acc.name
                    ledger_id = acc.code

            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=ledger_id)
            ws.cell(row=row, column=3, value=account_code)
            ws.cell(row=row, column=4, value=account_name)
            ws.cell(row=row, column=5, value=float(tot["cr_amount"]))
            ws.cell(row=row, column=6, value=float(tot["dr_amount"]))
            ws.cell(row=row, column=7, value=float(tot["total_amount"]))
            ws.cell(row=row, column=8, value=tot["entry_count"])
            for col in [5, 6, 7]:
                ws.cell(row=row, column=col).number_format = "#,##0.00"
            row += 1

        # Grand total row
        if by_year_account:
            grand_cr = sum(float(t["cr_amount"]) for t in by_year_account.values())
            grand_dr = sum(float(t["dr_amount"]) for t in by_year_account.values())
            grand_net = sum(float(t["total_amount"]) for t in by_year_account.values())
            grand_count = sum(t["entry_count"] for t in by_year_account.values())
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.merge_cells(f"A{row}:D{row}")
            ws.cell(row=row, column=5, value=grand_cr)
            ws.cell(row=row, column=6, value=grand_dr)
            ws.cell(row=row, column=7, value=grand_net)
            ws.cell(row=row, column=8, value=grand_count)
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.COLOR_TOTAL,
                    end_color=self.COLOR_TOTAL,
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color="FFFFFF")
                if col in [5, 6, 7]:
                    cell.number_format = "#,##0.00"

        for col_letter, width in [("A", 8), ("B", 12), ("C", 15), ("D", 40), ("E", 16), ("F", 16), ("G", 16), ("H", 12)]:
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = ws.dimensions

    def _add_account_summary_by_quarter_sheet(
        self,
        wb: Workbook,
        aggregations: list[QuarterlyAggregation],
        index: int,
    ) -> None:
        """One row per account per quarter per year: Year, Quarter, Ledger ID, Account Code, Account Name, CR, DR, Net, Entry Count."""
        ws = wb.create_sheet("Account Summary (by Quarter)", index)
        ws["A1"] = "Account Summary by Quarter"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        headers = [
            "Year",
            "Quarter",
            "Ledger ID",
            "Account Code",
            "Account Name",
            "CR Amount",
            "DR Amount",
            "Net Amount",
            "Entry Count",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
        row = 4

        for agg in sorted(aggregations, key=lambda a: (a.year, a.quarter, a.account_code)):
            ledger_id = agg.account_code
            account_name = agg.account_path
            if self.account_hierarchy:
                acc = self.account_hierarchy.get_account(agg.account_code)
                if not acc:
                    acc = self.account_hierarchy.get_account_by_name(agg.account_code)
                if acc:
                    account_name = acc.name
                    ledger_id = acc.code

            ws.cell(row=row, column=1, value=agg.year)
            ws.cell(row=row, column=2, value=agg.quarter)
            ws.cell(row=row, column=3, value=ledger_id)
            ws.cell(row=row, column=4, value=agg.account_code)
            ws.cell(row=row, column=5, value=account_name)
            ws.cell(row=row, column=6, value=float(agg.cr_amount))
            ws.cell(row=row, column=7, value=float(agg.dr_amount))
            ws.cell(row=row, column=8, value=float(agg.total_amount))
            ws.cell(row=row, column=9, value=agg.entry_count)
            for col in [6, 7, 8]:
                ws.cell(row=row, column=col).number_format = "#,##0.00"
            row += 1

        for col_letter, width in [("A", 8), ("B", 8), ("C", 12), ("D", 15), ("E", 40), ("F", 16), ("G", 16), ("H", 16), ("I", 12)]:
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = ws.dimensions

    def _add_audit_review_sheet(
        self,
        wb: Workbook,
        audit_trail: AuditTrail,
        error_detector: Optional[ErrorDetector],
        journal_entries: list[JournalEntry],
        ledger_entries: list[LedgerEntry],
        index: int,
    ) -> None:
        """Single sheet: summary stats, unmatched entries, and optional error counts."""
        ws = wb.create_sheet("Audit & Review", index)
        ws["A1"] = "Audit & Review"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        row = 3
        summary = audit_trail.get_summary()
        unmatched = audit_trail.get_unmatched_entries()

        # Summary
        ws.cell(row=row, column=1, value="Summary")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        if error_detector:
            ws.cell(row=row, column=1, value="Errors (validation)")
            ws.cell(row=row, column=2, value=len(error_detector.get_all_errors()))
            row += 1
            ws.cell(row=row, column=1, value="Warnings (validation)")
            ws.cell(row=row, column=2, value=len(error_detector.get_all_warnings()))
            row += 1
        row += 1

        # Unmatched entries (need review)
        ws.cell(row=row, column=1, value="Unmatched journal entries (no mapping rule)")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        headers = ["Entry ID", "Description", "Date", "Amount", "Year"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
        row += 1
        for entry in unmatched[:500]:  # Cap for sheet size
            ws.cell(row=row, column=1, value=entry.entry_id)
            ws.cell(row=row, column=2, value=entry.description)
            ws.cell(row=row, column=3, value=entry.date.isoformat() if entry.date else "")
            ws.cell(row=row, column=4, value=float(entry.amount))
            ws.cell(row=row, column=5, value=entry.year)
            row += 1
        if len(unmatched) > 500:
            ws.cell(row=row, column=1, value=f"... and {len(unmatched) - 500} more")
            row += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 8
        ws.freeze_panes = "A2"
