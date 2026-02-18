"""Excel formatting utilities for consistent styling across all reports."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """
    Provides consistent Excel formatting across all output files.

    Centralizes formatting styles, colors, and formatting patterns
    for professional, consistent output.
    """

    # Color palette
    COLOR_HEADER = "2C3E50"  # Dark blue-gray
    COLOR_TOTAL = "3498DB"  # Blue
    COLOR_SUCCESS = "2ECC71"  # Green
    COLOR_WARNING = "F39C12"  # Orange
    COLOR_ERROR = "E74C3C"  # Red
    COLOR_CRITICAL = "8B0000"  # Dark red
    COLOR_INFO = "95A5A6"  # Gray

    # Level colors for hierarchy
    COLOR_LEVEL1 = "34495E"  # Medium gray
    COLOR_LEVEL2 = "7F8C8D"  # Light gray
    COLOR_LEVEL3 = "BDC3C7"  # Very light gray
    COLOR_LEVEL4 = "ECF0F1"  # Almost white

    # Quarter colors
    COLOR_QUARTER1 = "E74C3C"  # Red
    COLOR_QUARTER2 = "F39C12"  # Orange
    COLOR_QUARTER3 = "2ECC71"  # Green
    COLOR_QUARTER4 = "9B59B6"  # Purple

    @staticmethod
    def apply_header_style(cell, text: str = None) -> None:
        """
        Apply header cell styling.

        Args:
            cell: OpenPyXL cell object
            text: Optional text to set in cell
        """
        if text:
            cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color=ExcelFormatter.COLOR_HEADER,
            end_color=ExcelFormatter.COLOR_HEADER,
            fill_type="solid",
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    @staticmethod
    def apply_total_style(cell, text: str = None) -> None:
        """
        Apply total row styling.

        Args:
            cell: OpenPyXL cell object
            text: Optional text to set in cell
        """
        if text:
            cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color=ExcelFormatter.COLOR_TOTAL,
            end_color=ExcelFormatter.COLOR_TOTAL,
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def apply_level_style(cell, level: int) -> None:
        """
        Apply hierarchy level styling.

        Args:
            cell: OpenPyXL cell object
            level: Hierarchy level (1-4)
        """
        colors = {
            1: ExcelFormatter.COLOR_LEVEL1,
            2: ExcelFormatter.COLOR_LEVEL2,
            3: ExcelFormatter.COLOR_LEVEL3,
            4: ExcelFormatter.COLOR_LEVEL4,
        }
        color = colors.get(level)
        if color:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )

    @staticmethod
    def apply_quarter_style(cell, quarter: int) -> None:
        """
        Apply quarter styling.

        Args:
            cell: OpenPyXL cell object
            quarter: Quarter number (1-4)
        """
        colors = {
            1: ExcelFormatter.COLOR_QUARTER1,
            2: ExcelFormatter.COLOR_QUARTER2,
            3: ExcelFormatter.COLOR_QUARTER3,
            4: ExcelFormatter.COLOR_QUARTER4,
        }
        color = colors.get(quarter)
        if color:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )

    @staticmethod
    def apply_severity_style(cell, severity: str) -> None:
        """
        Apply severity-based styling.

        Args:
            cell: OpenPyXL cell object
            severity: Severity level (critical, error, warning, info)
        """
        colors = {
            "critical": ExcelFormatter.COLOR_CRITICAL,
            "error": ExcelFormatter.COLOR_ERROR,
            "warning": ExcelFormatter.COLOR_WARNING,
            "info": ExcelFormatter.COLOR_INFO,
            "success": ExcelFormatter.COLOR_SUCCESS,
        }
        color = colors.get(severity.lower())
        if color:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )
            if severity.lower() in ["critical", "error"]:
                cell.font = Font(bold=True, color="FFFFFF")
            else:
                cell.font = Font(bold=True)

    @staticmethod
    def format_amount_cell(cell, value: float | None = None) -> None:
        """
        Format cell as currency/amount.

        Args:
            cell: OpenPyXL cell object
            value: Optional amount value to set
        """
        if value is not None:
            cell.value = value
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def format_percentage_cell(cell, value: float | None = None) -> None:
        """
        Format cell as percentage.

        Args:
            cell: OpenPyXL cell object
            value: Optional percentage value to set (0.0-1.0 or 0-100)
        """
        if value is not None:
            # If value > 1, assume it's already a percentage (0-100)
            if value > 1:
                cell.value = value / 100
            else:
                cell.value = value
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def format_date_cell(cell, value: str | None = None) -> None:
        """
        Format cell as date.

        Args:
            cell: OpenPyXL cell object
            value: Optional date value to set (ISO format string)
        """
        if value:
            cell.value = value
        cell.number_format = "YYYY-MM-DD"
        cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def apply_border(cell) -> None:
        """
        Apply standard border to cell.

        Args:
            cell: OpenPyXL cell object
        """
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    @staticmethod
    def auto_adjust_column_widths(ws, column_widths: dict[str, int] | None = None) -> None:
        """
        Auto-adjust column widths for worksheet.

        Args:
            ws: OpenPyXL worksheet object
            column_widths: Optional dictionary mapping column letters to widths
        """
        if column_widths:
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            # Auto-adjust all columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def setup_filters(ws, header_row: int = 1) -> None:
        """
        Setup auto-filters for worksheet.

        Args:
            ws: OpenPyXL worksheet object
            header_row: Row number containing headers (default: 1)
        """
        if ws.max_row > header_row:
            ws.auto_filter.ref = ws.dimensions

    @staticmethod
    def freeze_panes(ws, freeze_cell: str = "A2") -> None:
        """
        Freeze panes at specified cell.

        Args:
            ws: OpenPyXL worksheet object
            freeze_cell: Cell reference to freeze at (default: "A2")
        """
        ws.freeze_panes = freeze_cell








