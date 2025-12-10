"""Excel error report generation for veritas-accounting."""

from collections import defaultdict
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from veritas_accounting.audit.trail import AuditTrail
from veritas_accounting.models.journal import JournalEntry
from veritas_accounting.validation.error_detector import DetectedError, ErrorDetector
from veritas_accounting.validation.validation_results import ValidationSummary
from veritas_accounting.utils.exceptions import ExcelIOError


class ErrorReportGenerator:
    """
    Generates comprehensive Excel error reports.

    Creates multi-sheet Excel files with error summaries, detailed errors,
    transformation logs, validation results, and auto-fix suggestions.
    """

    # Color scheme for severity levels
    COLOR_CRITICAL = "FF0000"  # Red
    COLOR_ERROR = "FF6B6B"  # Light red
    COLOR_WARNING = "FFD93D"  # Yellow
    COLOR_INFO = "6BCB77"  # Green
    COLOR_SUCCESS = "4ECDC4"  # Teal
    COLOR_HEADER = "2C3E50"  # Dark blue-gray

    def __init__(self) -> None:
        """Initialize ErrorReportGenerator."""
        pass

    def generate_report(
        self,
        output_path: Path | str,
        error_detector: ErrorDetector,
        audit_trail: Optional[AuditTrail] = None,
        validation_summary: Optional[ValidationSummary] = None,
        original_entries: Optional[list[JournalEntry]] = None,
        auto_fix_suggestions: Optional[list[dict[str, Any]]] = None,
    ) -> None:
        """
        Generate comprehensive error report Excel file.

        Args:
            output_path: Path to output Excel file
            error_detector: ErrorDetector with all errors and warnings
            audit_trail: Optional AuditTrail with transformation records
            validation_summary: Optional ValidationSummary with validation results
            original_entries: Optional list of original journal entries for comparison
            auto_fix_suggestions: Optional list of auto-fix suggestions

        Raises:
            ExcelIOError: If report generation fails
        """
        output_path = Path(output_path)

        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Generate all sheets
            self._create_summary_sheet(
                wb, error_detector, validation_summary, audit_trail
            )
            self._create_errors_sheet(wb, error_detector)
            if audit_trail:
                self._create_transformations_sheet(wb, audit_trail)
            if validation_summary:
                self._create_validation_sheet(wb, validation_summary)
            if auto_fix_suggestions:
                self._create_auto_fixes_sheet(wb, auto_fix_suggestions)
            if original_entries:
                self._create_original_data_sheet(wb, original_entries)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

        except Exception as e:
            raise ExcelIOError(
                f"Failed to generate error report: {output_path}. Error: {e}"
            ) from e

    def _create_summary_sheet(
        self,
        wb: Workbook,
        error_detector: ErrorDetector,
        validation_summary: Optional[ValidationSummary],
        audit_trail: Optional[AuditTrail],
    ) -> None:
        """Create summary sheet with overview statistics."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Error Report Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        row = 3

        # Overall Statistics
        ws[f"A{row}"] = "Overall Statistics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        all_errors = error_detector.get_all_errors()
        all_warnings = error_detector.get_all_warnings()

        stats = [
            ("Total Errors", len(all_errors)),
            ("Total Warnings", len(all_warnings)),
            ("Critical Errors", len(error_detector.get_errors_by_severity("critical"))),
            ("Data Errors", len(error_detector.get_errors_by_type("data_error"))),
            ("Rule Errors", len(error_detector.get_errors_by_type("rule_error"))),
            (
                "Transformation Errors",
                len(error_detector.get_errors_by_type("transformation_error")),
            ),
            ("Output Errors", len(error_detector.get_errors_by_type("output_error"))),
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            # Color code based on value
            if "Error" in label and value > 0:
                ws[f"B{row}"].fill = PatternFill(
                    start_color=self.COLOR_ERROR, end_color=self.COLOR_ERROR, fill_type="solid"
                )
            row += 1

        # Validation Summary
        if validation_summary:
            row += 1
            ws[f"A{row}"] = "Validation Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            val_stats = [
                ("Status", validation_summary.overall_status.upper()),
                ("Entries Processed", validation_summary.entries_processed),
                ("Valid Entries", validation_summary.entries_valid),
                ("Errors Found", validation_summary.errors_found),
                ("Warnings Found", validation_summary.warnings_found),
                (
                    "Overall Confidence",
                    f"{validation_summary.overall_confidence.level.upper()} ({validation_summary.overall_confidence.score:.1%})",
                ),
            ]

            for label, value in val_stats:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value
                # Color code status
                if label == "Status":
                    if validation_summary.status_indicator == "green":
                        ws[f"B{row}"].fill = PatternFill(
                            start_color=self.COLOR_SUCCESS,
                            end_color=self.COLOR_SUCCESS,
                            fill_type="solid",
                        )
                    elif validation_summary.status_indicator == "yellow":
                        ws[f"B{row}"].fill = PatternFill(
                            start_color=self.COLOR_WARNING,
                            end_color=self.COLOR_WARNING,
                            fill_type="solid",
                        )
                    else:
                        ws[f"B{row}"].fill = PatternFill(
                            start_color=self.COLOR_ERROR,
                            end_color=self.COLOR_ERROR,
                            fill_type="solid",
                        )
                row += 1

        # Audit Trail Summary
        if audit_trail:
            row += 1
            ws[f"A{row}"] = "Transformation Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            summary = audit_trail.get_summary()
            trail_stats = [
                ("Total Transformations", summary["total_transformations"]),
                ("Matched Entries", summary["matched_entries"]),
                ("Unmatched Entries", summary["unmatched_entries"]),
                ("Total Ledger Entries", summary["total_ledger_entries"]),
                ("Unique Rules Applied", summary["unique_rules_applied"]),
            ]

            for label, value in trail_stats:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value
                row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_errors_sheet(self, wb: Workbook, error_detector: ErrorDetector) -> None:
        """Create detailed errors sheet."""
        ws = wb.create_sheet("Errors")

        # Headers
        headers = [
            "Row Number",
            "Entry ID",
            "Rule ID",
            "Field Name",
            "Error Type",
            "Severity",
            "Error Message",
            "Actual Value",
            "Expected Value",
            "Requires Review",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Error data
        all_errors = error_detector.get_all_errors()
        for row_idx, error in enumerate(all_errors, start=2):
            ws.cell(row=row_idx, column=1, value=error.row_number)
            ws.cell(row=row_idx, column=2, value=error.entry_id or "")
            ws.cell(row=row_idx, column=3, value=error.rule_id or "")
            ws.cell(row=row_idx, column=4, value=error.field_name)
            ws.cell(row=row_idx, column=5, value=error.error_type)
            ws.cell(row=row_idx, column=6, value=error.severity)
            ws.cell(row=row_idx, column=7, value=error.error_message)
            ws.cell(row=row_idx, column=8, value=str(error.actual_value))
            ws.cell(row=row_idx, column=9, value=str(error.expected_value) if error.expected_value else "")
            ws.cell(row=row_idx, column=10, value="Yes" if error.requires_review else "No")

            # Color code by severity
            severity_colors = {
                "critical": self.COLOR_CRITICAL,
                "error": self.COLOR_ERROR,
                "warning": self.COLOR_WARNING,
                "info": self.COLOR_INFO,
            }
            if error.severity in severity_colors:
                ws.cell(row=row_idx, column=6).fill = PatternFill(
                    start_color=severity_colors[error.severity],
                    end_color=severity_colors[error.severity],
                    fill_type="solid",
                )

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_transformations_sheet(
        self, wb: Workbook, audit_trail: AuditTrail
    ) -> None:
        """Create transformations log sheet."""
        ws = wb.create_sheet("Transformations")

        # Headers
        headers = [
            "Entry ID",
            "Timestamp",
            "Source Description",
            "Source Amount",
            "Source Date",
            "Applied Rules",
            "Generated Ledger Entries",
            "No Match",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Transformation data
        for row_idx, record in enumerate(audit_trail.records, start=2):
            ws.cell(row=row_idx, column=1, value=record.entry_id)
            ws.cell(row=row_idx, column=2, value=record.timestamp.isoformat())
            ws.cell(row=row_idx, column=3, value=record.source_entry.description)
            ws.cell(row=row_idx, column=4, value=float(record.source_entry.amount))
            ws.cell(row=row_idx, column=5, value=record.source_entry.date.isoformat())
            ws.cell(
                row=row_idx,
                column=6,
                value=", ".join(rule.rule_id for rule in record.applied_rules),
            )
            ws.cell(row=row_idx, column=7, value=len(record.generated_entries))
            ws.cell(row=row_idx, column=8, value="Yes" if record.no_match else "No")

            # Color code no-match entries
            if record.no_match:
                ws.cell(row=row_idx, column=8).fill = PatternFill(
                    start_color=self.COLOR_WARNING,
                    end_color=self.COLOR_WARNING,
                    fill_type="solid",
                )

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_validation_sheet(
        self, wb: Workbook, validation_summary: ValidationSummary
    ) -> None:
        """Create validation results sheet."""
        ws = wb.create_sheet("Validation")

        # Headers
        headers = [
            "Metric",
            "Value",
            "Status",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Validation data
        row = 2
        metrics = [
            ("Overall Status", validation_summary.overall_status.upper(), validation_summary.status_indicator),
            ("Entries Processed", validation_summary.entries_processed, None),
            ("Valid Entries", validation_summary.entries_valid, None),
            ("Errors Found", validation_summary.errors_found, None),
            ("Warnings Found", validation_summary.warnings_found, None),
            ("Error Rate", f"{validation_summary.error_rate:.2%}", None),
            (
                "Overall Confidence",
                f"{validation_summary.overall_confidence.level.upper()} ({validation_summary.overall_confidence.score:.1%})",
                validation_summary.overall_confidence.level,
            ),
        ]

        for metric, value, status in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            if status:
                ws.cell(row=row, column=3, value=status)
                # Color code status
                if status == "green" or status == "high":
                    ws.cell(row=row, column=3).fill = PatternFill(
                        start_color=self.COLOR_SUCCESS,
                        end_color=self.COLOR_SUCCESS,
                        fill_type="solid",
                    )
                elif status == "yellow" or status == "medium":
                    ws.cell(row=row, column=3).fill = PatternFill(
                        start_color=self.COLOR_WARNING,
                        end_color=self.COLOR_WARNING,
                        fill_type="solid",
                    )
                elif status == "red" or status == "low":
                    ws.cell(row=row, column=3).fill = PatternFill(
                        start_color=self.COLOR_ERROR,
                        end_color=self.COLOR_ERROR,
                        fill_type="solid",
                    )
            row += 1

        # Coverage section
        row += 1
        ws.cell(row=row, column=1, value="Validation Coverage")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for stage, covered in validation_summary.coverage.items():
            ws.cell(row=row, column=1, value=stage.title())
            ws.cell(row=row, column=2, value="✅ Covered" if covered else "⏭️ Not validated")
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15

    def _create_auto_fixes_sheet(
        self, wb: Workbook, auto_fix_suggestions: list[dict[str, Any]]
    ) -> None:
        """Create auto-fix suggestions sheet."""
        ws = wb.create_sheet("Auto-Fixes")

        # Headers
        headers = [
            "Entry ID",
            "Field Name",
            "Original Value",
            "Suggested Value",
            "Confidence",
            "Approval Status",
            "Fix Description",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Auto-fix data
        for row_idx, suggestion in enumerate(auto_fix_suggestions, start=2):
            ws.cell(row=row_idx, column=1, value=suggestion.get("entry_id", ""))
            ws.cell(row=row_idx, column=2, value=suggestion.get("field_name", ""))
            ws.cell(row=row_idx, column=3, value=str(suggestion.get("original_value", "")))
            ws.cell(row=row_idx, column=4, value=str(suggestion.get("suggested_value", "")))
            ws.cell(row=row_idx, column=5, value=suggestion.get("confidence", ""))
            ws.cell(row=row_idx, column=6, value=suggestion.get("approval_status", "Pending"))
            ws.cell(row=row_idx, column=7, value=suggestion.get("description", ""))

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_original_data_sheet(
        self, wb: Workbook, original_entries: list[JournalEntry]
    ) -> None:
        """Create original data sheet for comparison."""
        ws = wb.create_sheet("Original Data")

        # Headers
        headers = [
            "Entry ID",
            "Year",
            "Description",
            "Old Type",
            "Amount",
            "Date",
            "Quarter",
            "Notes",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")

        # Original entry data
        for row_idx, entry in enumerate(original_entries, start=2):
            ws.cell(row=row_idx, column=1, value=entry.entry_id)
            ws.cell(row=row_idx, column=2, value=entry.year)
            ws.cell(row=row_idx, column=3, value=entry.description)
            ws.cell(row=row_idx, column=4, value=entry.old_type)
            ws.cell(row=row_idx, column=5, value=float(entry.amount))
            ws.cell(row=row_idx, column=6, value=entry.date.isoformat())
            ws.cell(row=row_idx, column=7, value=entry.quarter or "")
            ws.cell(row=row_idx, column=8, value=entry.notes or "")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions
