"""Ledger output Excel generation for veritas-accounting."""

from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from veritas_accounting.models.account import AccountHierarchy
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.utils.exceptions import ExcelIOError


class LedgerOutputGenerator:
    """
    Generates ledger output in Excel format with hierarchical organization.

    Creates professional Excel files with ledger entries organized by account
    hierarchy, proper formatting, and summary totals.
    """

    # Color scheme
    COLOR_HEADER = "2C3E50"  # Dark blue-gray
    COLOR_LEVEL1 = "34495E"  # Medium gray
    COLOR_LEVEL2 = "7F8C8D"  # Light gray
    COLOR_LEVEL3 = "BDC3C7"  # Very light gray
    COLOR_LEVEL4 = "ECF0F1"  # Almost white
    COLOR_TOTAL = "3498DB"  # Blue

    def __init__(
        self, account_hierarchy: AccountHierarchy | None = None
    ) -> None:
        """
        Initialize LedgerOutputGenerator.

        Args:
            account_hierarchy: AccountHierarchy object (optional, for hierarchical organization)
        """
        self.account_hierarchy = account_hierarchy

    def generate(
        self,
        ledger_entries: list[LedgerEntry],
        output_path: Path | str,
    ) -> None:
        """
        Generate ledger output Excel file.

        Args:
            ledger_entries: List of LedgerEntry objects
            output_path: Path to output Excel file

        Raises:
            ExcelIOError: If generation fails
        """
        output_path = Path(output_path)

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create sheets
            self._create_ledger_entries_sheet(wb, ledger_entries)
            self._create_summary_sheet(wb, ledger_entries)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

        except Exception as e:
            raise ExcelIOError(
                f"Failed to generate ledger output: {output_path}. Error: {e}"
            ) from e

    def _create_ledger_entries_sheet(
        self, wb: Workbook, ledger_entries: list[LedgerEntry]
    ) -> None:
        """Create ledger entries sheet with hierarchical organization."""
        ws = wb.create_sheet("Ledger Entries", 0)

        # Headers
        headers = [
            "Entry ID",
            "Account Code",
            "Account Path",
            "Level",
            "Description",
            "Amount",
            "Date",
            "Quarter",
            "Year",
            "Source Entry ID",
            "Rule Applied",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Organize entries by account hierarchy
        if self.account_hierarchy:
            # Group by account code and level
            organized_entries = self._organize_by_hierarchy(ledger_entries)
        else:
            # Simple sort by account code
            organized_entries = sorted(
                ledger_entries, key=lambda e: (e.account_code, e.date)
            )

        # Write entries
        row = 2
        for entry in organized_entries:
            level = self._get_account_level(entry.account_code)

            ws.cell(row=row, column=1, value=entry.entry_id)
            ws.cell(row=row, column=2, value=entry.account_code)
            ws.cell(row=row, column=3, value=entry.account_path)
            ws.cell(row=row, column=4, value=level)
            ws.cell(row=row, column=5, value=entry.description)
            ws.cell(row=row, column=6, value=float(entry.amount))
            ws.cell(row=row, column=7, value=entry.date.isoformat())
            ws.cell(row=row, column=8, value=entry.quarter)
            ws.cell(row=row, column=9, value=entry.year)
            ws.cell(row=row, column=10, value=entry.source_entry_id)
            ws.cell(row=row, column=11, value=entry.rule_applied)

            # Apply level-based formatting
            level_color = self._get_level_color(level)
            if level_color:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(
                        start_color=level_color,
                        end_color=level_color,
                        fill_type="solid",
                    )

            # Format amount column
            amount_cell = ws.cell(row=row, column=6)
            amount_cell.number_format = "#,##0.00"

            # Format date column
            date_cell = ws.cell(row=row, column=7)
            date_cell.number_format = "YYYY-MM-DD"

            row += 1

        # Auto-adjust column widths
        column_widths = {
            "A": 15,  # Entry ID
            "B": 15,  # Account Code
            "C": 40,  # Account Path
            "D": 8,   # Level
            "E": 30,  # Description
            "F": 15,  # Amount
            "G": 12,  # Date
            "H": 8,   # Quarter
            "I": 8,   # Year
            "J": 15,  # Source Entry ID
            "K": 15,  # Rule Applied
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _create_summary_sheet(
        self, wb: Workbook, ledger_entries: list[LedgerEntry]
    ) -> None:
        """Create summary sheet with hierarchical totals."""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Hierarchical Summary Totals"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        if not self.account_hierarchy:
            ws["A3"] = "No account hierarchy available for summary"
            return

        # Calculate totals by level
        level_totals = self._calculate_level_totals(ledger_entries)

        row = 3
        for level in [1, 2, 3, 4]:
            if level not in level_totals:
                continue

            # Level header
            ws[f"A{row}"] = f"Level {level} Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:D{row}")
            row += 1

            # Sub-headers
            headers = ["Account Code", "Account Path", "Total Amount", "Entry Count"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.COLOR_HEADER,
                    end_color=self.COLOR_HEADER,
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            row += 1

            # Level totals
            for account_code, totals in sorted(level_totals[level].items()):
                account = self.account_hierarchy.get_account(account_code)
                account_path = account.full_path if account else account_code

                ws.cell(row=row, column=1, value=account_code)
                ws.cell(row=row, column=2, value=account_path)
                ws.cell(row=row, column=3, value=float(totals["amount"]))
                ws.cell(row=row, column=4, value=totals["count"])

                # Format amount
                amount_cell = ws.cell(row=row, column=3)
                amount_cell.number_format = "#,##0.00"

                # Apply level color
                level_color = self._get_level_color(level)
                if level_color:
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(
                            start_color=level_color,
                            end_color=level_color,
                            fill_type="solid",
                        )

                row += 1

            # Level total row
            total_amount = sum(t["amount"] for t in level_totals[level].values())
            total_count = sum(t["count"] for t in level_totals[level].values())

            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"Level {level} Total")
            ws.cell(row=row, column=3, value=float(total_amount))
            ws.cell(row=row, column=4, value=total_count)

            # Format total row
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.COLOR_TOTAL,
                    end_color=self.COLOR_TOTAL,
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color="FFFFFF")
                if col == 3:
                    cell.number_format = "#,##0.00"

            row += 2  # Space between levels

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12

    def _organize_by_hierarchy(
        self, ledger_entries: list[LedgerEntry]
    ) -> list[LedgerEntry]:
        """
        Organize entries by account hierarchy (Level 1 → Level 4).

        Args:
            ledger_entries: List of LedgerEntry objects

        Returns:
            List of LedgerEntry objects organized by hierarchy
        """
        if not self.account_hierarchy:
            return sorted(ledger_entries, key=lambda e: (e.account_code, e.date))

        # Group by level, then by account code
        by_level: dict[int, dict[str, list[LedgerEntry]]] = defaultdict(
            lambda: defaultdict(list)
        )

        for entry in ledger_entries:
            level = self._get_account_level(entry.account_code)
            by_level[level][entry.account_code].append(entry)

        # Sort entries within each account by date
        for level_dict in by_level.values():
            for account_code in level_dict:
                level_dict[account_code].sort(key=lambda e: e.date)

        # Build ordered list: Level 1 → Level 4
        organized: list[LedgerEntry] = []
        for level in [1, 2, 3, 4]:
            if level in by_level:
                # Sort accounts within level
                sorted_accounts = sorted(by_level[level].keys())
                for account_code in sorted_accounts:
                    organized.extend(by_level[level][account_code])

        return organized

    def _calculate_level_totals(
        self, ledger_entries: list[LedgerEntry]
    ) -> dict[int, dict[str, dict[str, Any]]]:
        """
        Calculate totals by level and account.

        Args:
            ledger_entries: List of LedgerEntry objects

        Returns:
            Dictionary mapping level to account_code to totals dict
        """
        level_totals: dict[int, dict[str, dict[str, Any]]] = defaultdict(
            lambda: defaultdict(lambda: {"amount": 0.0, "count": 0})
        )

        for entry in ledger_entries:
            level = self._get_account_level(entry.account_code)
            if level > 0:
                level_totals[level][entry.account_code]["amount"] += float(
                    entry.amount
                )
                level_totals[level][entry.account_code]["count"] += 1

        return level_totals

    def _get_account_level(self, account_code: str) -> int:
        """
        Get account level from hierarchy.

        Args:
            account_code: Account code

        Returns:
            Account level (1-4) or 0 if not found
        """
        if self.account_hierarchy:
            account = self.account_hierarchy.get_account(account_code)
            if account:
                return account.level
        return 0

    def _get_level_color(self, level: int) -> str | None:
        """
        Get color for hierarchy level.

        Args:
            level: Hierarchy level (1-4)

        Returns:
            Color hex code or None
        """
        colors = {
            1: self.COLOR_LEVEL1,
            2: self.COLOR_LEVEL2,
            3: self.COLOR_LEVEL3,
            4: self.COLOR_LEVEL4,
        }
        return colors.get(level)
