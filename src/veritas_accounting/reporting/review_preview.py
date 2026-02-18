"""Review preview report generation for veritas-accounting.

This module provides a user-friendly preview and review system that shows
accountants the transformed ledger entries in a table format with visual
flags for problematic entries that need review.
"""

from collections import defaultdict
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from veritas_accounting.audit.trail import AuditTrail
from veritas_accounting.models.journal import JournalEntry
from veritas_accounting.models.ledger import LedgerEntry
from veritas_accounting.validation.error_detector import DetectedError, ErrorDetector
from veritas_accounting.utils.exceptions import ExcelIOError


class ReviewStatus(Enum):
    """Status flags for entries needing review."""

    OK = "ok"  # No issues
    NO_MATCH = "no_match"  # No matching rule found
    VALIDATION_ERROR = "validation_error"  # Validation error
    VALIDATION_WARNING = "validation_warning"  # Validation warning
    UNUSUAL_AMOUNT = "unusual_amount"  # Amount is unusually large/small
    MISSING_ACCOUNT = "missing_account"  # Account code not found in hierarchy
    MULTIPLE_RULES = "multiple_rules"  # Multiple rules matched (potential conflict)


@dataclass
class EntryReviewFlag:
    """Flag for an entry that needs review."""

    entry_id: str
    status: ReviewStatus
    reason: str
    severity: str  # "critical", "error", "warning", "info"
    source_entry_id: Optional[str] = None


class ReviewPreviewGenerator:
    """
    Generates review preview reports for accountants.

    Creates Excel files with:
    - Preview table of all ledger entries (the final output)
    - Visual flags for problematic entries
    - Side-by-side comparison (Journal → Ledger)
    - Review dashboard with summary of issues
    """

    # Color scheme
    COLOR_HEADER = "2C3E50"  # Dark blue-gray
    COLOR_OK = "E8F5E9"  # Light green (no issues)
    COLOR_CRITICAL = "FFEBEE"  # Light red (critical issues)
    COLOR_ERROR = "FFF3E0"  # Light orange (errors)
    COLOR_WARNING = "FFFDE7"  # Light yellow (warnings)
    COLOR_INFO = "E3F2FD"  # Light blue (info)
    COLOR_BORDER = "CCCCCC"  # Gray border

    # Status icons (using Unicode symbols)
    ICON_OK = "✓"
    ICON_ERROR = "⚠"
    ICON_CRITICAL = "✗"
    ICON_WARNING = "!"
    ICON_INFO = "ℹ"

    def __init__(
        self,
        account_hierarchy: Optional[Any] = None,
        error_detector: Optional[ErrorDetector] = None,
        audit_trail: Optional[AuditTrail] = None,
    ) -> None:
        """
        Initialize ReviewPreviewGenerator.

        Args:
            account_hierarchy: AccountHierarchy object (optional)
            error_detector: ErrorDetector with validation errors (optional)
            audit_trail: AuditTrail with transformation records (optional)
        """
        self.account_hierarchy = account_hierarchy
        self.error_detector = error_detector
        self.audit_trail = audit_trail
        self.review_flags: list[EntryReviewFlag] = []

    def generate(
        self,
        ledger_entries: list[LedgerEntry],
        journal_entries: list[JournalEntry],
        output_path: Path | str,
    ) -> None:
        """
        Generate review preview Excel file.

        Args:
            ledger_entries: List of transformed LedgerEntry objects
            journal_entries: List of original JournalEntry objects
            output_path: Path to output Excel file

        Raises:
            ExcelIOError: If generation fails
        """
        output_path = Path(output_path)

        try:
            # Analyze entries and create review flags
            self._analyze_entries(ledger_entries, journal_entries)

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create sheets in order
            self._create_review_dashboard(wb, ledger_entries, journal_entries)
            self._create_preview_table(wb, ledger_entries)
            self._create_comparison_view(wb, ledger_entries, journal_entries)
            self._create_flagged_entries(wb, ledger_entries, journal_entries)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

        except Exception as e:
            raise ExcelIOError(
                f"Failed to generate review preview: {output_path}. Error: {e}"
            ) from e

    def _analyze_entries(
        self,
        ledger_entries: list[LedgerEntry],
        journal_entries: list[JournalEntry],
    ) -> None:
        """Analyze entries and create review flags."""
        self.review_flags = []

        # Create lookup maps
        journal_by_id = {entry.entry_id: entry for entry in journal_entries}
        ledger_by_source = defaultdict(list)
        for entry in ledger_entries:
            if entry.source_entry_id:
                ledger_by_source[entry.source_entry_id].append(entry)

        # Check for entries with no matching rules (no_match)
        if self.audit_trail:
            for record in self.audit_trail.records:
                if record.no_match:
                    self.review_flags.append(
                        EntryReviewFlag(
                            entry_id=record.entry_id,
                            status=ReviewStatus.NO_MATCH,
                            reason="No matching rule found for this journal entry",
                            severity="error",
                            source_entry_id=record.entry_id,
                        )
                    )

        # Check validation errors
        # Errors reference journal entry IDs, not ledger entry IDs
        # We need to map them to ledger entries or flag the journal entry itself
        if self.error_detector:
            # Create map of journal entry IDs to ledger entry IDs
            journal_to_ledger_map: dict[str, list[str]] = defaultdict(list)
            for ledger_entry in ledger_entries:
                if ledger_entry.source_entry_id:
                    journal_to_ledger_map[ledger_entry.source_entry_id].append(
                        ledger_entry.entry_id
                    )

            # Process errors - they may reference journal entries (source) or ledger entries
            for error in self.error_detector.get_all_errors():
                if not error.entry_id:
                    # Error without entry_id - flag the first ledger entry or create a generic flag
                    if ledger_entries:
                        first_entry = ledger_entries[0]
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=first_entry.entry_id,
                                status=ReviewStatus.VALIDATION_ERROR,
                                reason=f"{error.error_message} (Row {error.row_number}, Field: {error.field_name})",
                                severity=error.severity,
                                source_entry_id=None,
                            )
                        )
                    continue

                # Check if this entry_id is a journal entry (has corresponding ledger entries)
                if error.entry_id in journal_to_ledger_map:
                    # Journal entry error - flag all corresponding ledger entries
                    for ledger_entry_id in journal_to_ledger_map[error.entry_id]:
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=ledger_entry_id,
                                status=ReviewStatus.VALIDATION_ERROR,
                                reason=f"{error.error_message} (Row {error.row_number}, Field: {error.field_name})",
                                severity=error.severity,
                                source_entry_id=error.entry_id,
                            )
                        )
                else:
                    # Might be a ledger entry ID or orphaned error - flag it directly
                    # Check if it's a ledger entry
                    ledger_entry = next(
                        (e for e in ledger_entries if e.entry_id == error.entry_id), None
                    )
                    if ledger_entry:
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=error.entry_id,
                                status=ReviewStatus.VALIDATION_ERROR,
                                reason=f"{error.error_message} (Row {error.row_number}, Field: {error.field_name})",
                                severity=error.severity,
                                source_entry_id=ledger_entry.source_entry_id,
                            )
                        )
                    else:
                        # Journal entry with no ledger entries (orphaned error)
                        # Create a flag with the journal entry ID as the entry_id
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=error.entry_id,
                                status=ReviewStatus.VALIDATION_ERROR,
                                reason=f"{error.error_message} (Row {error.row_number}, Field: {error.field_name})",
                                severity=error.severity,
                                source_entry_id=error.entry_id,
                            )
                        )

        # Check validation warnings
        if self.error_detector:
            # Map journal entry IDs to ledger entry IDs for warnings too
            journal_to_ledger_map: dict[str, list[str]] = defaultdict(list)
            for ledger_entry in ledger_entries:
                if ledger_entry.source_entry_id:
                    journal_to_ledger_map[ledger_entry.source_entry_id].append(
                        ledger_entry.entry_id
                    )

            for warning in self.error_detector.get_all_warnings():
                if not warning.entry_id:
                    continue

                # Get warning message (ValidationWarning uses warning_message, not message)
                warning_msg = warning.warning_message if hasattr(warning, 'warning_message') else str(warning)

                # Map journal entry warnings to ledger entries
                if warning.entry_id in journal_to_ledger_map:
                    for ledger_entry_id in journal_to_ledger_map[warning.entry_id]:
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=ledger_entry_id,
                                status=ReviewStatus.VALIDATION_WARNING,
                                reason=warning_msg,
                                severity="warning",
                                source_entry_id=warning.entry_id,
                            )
                        )
                else:
                    # Direct ledger entry warning or orphaned
                    ledger_entry = next(
                        (e for e in ledger_entries if e.entry_id == warning.entry_id), None
                    )
                    if ledger_entry:
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=warning.entry_id,
                                status=ReviewStatus.VALIDATION_WARNING,
                                reason=warning_msg,
                                severity="warning",
                                source_entry_id=ledger_entry.source_entry_id,
                            )
                        )
                    else:
                        self.review_flags.append(
                            EntryReviewFlag(
                                entry_id=warning.entry_id,
                                status=ReviewStatus.VALIDATION_WARNING,
                                reason=warning_msg,
                                severity="warning",
                                source_entry_id=warning.entry_id,
                            )
                        )

        # Check for missing accounts (skip null/None account codes)
        for entry in ledger_entries:
            if not entry.account_code:  # Skip null/None account codes
                continue
            if self.account_hierarchy:
                account = self.account_hierarchy.get_account(entry.account_code)
                if not account:
                    # Try by name
                    account = self.account_hierarchy.get_account_by_name(
                        entry.account_code
                    )
                if not account:
                    self.review_flags.append(
                        EntryReviewFlag(
                            entry_id=entry.entry_id,
                            status=ReviewStatus.MISSING_ACCOUNT,
                            reason=f"Account code '{entry.account_code}' not found in hierarchy",
                            severity="warning",
                            source_entry_id=entry.source_entry_id,
                        )
                    )

        # Check for unusual amounts (statistical outliers)
        # Use median-based detection for better robustness with small datasets
        if ledger_entries and len(ledger_entries) >= 3:  # Need at least 3 entries for meaningful stats
            amounts = [float(entry.amount) for entry in ledger_entries]
            if amounts:
                # Use median and median absolute deviation (MAD) for robustness
                sorted_amounts = sorted(amounts)
                n = len(sorted_amounts)
                if n % 2 == 0:
                    median_amount = (sorted_amounts[n // 2 - 1] + sorted_amounts[n // 2]) / 2
                else:
                    median_amount = sorted_amounts[n // 2]
                
                # Calculate median absolute deviation (MAD)
                deviations = [abs(x - median_amount) for x in amounts]
                sorted_deviations = sorted(deviations)
                if n % 2 == 0:
                    mad = (sorted_deviations[n // 2 - 1] + sorted_deviations[n // 2]) / 2
                else:
                    mad = sorted_deviations[n // 2]
                
                # Use 5 * MAD as threshold (similar to 5 standard deviations, more conservative)
                # If MAD is 0 (all amounts are the same), skip detection
                if mad > 0:
                    threshold_high = median_amount + (5 * mad)  # Upper threshold
                    threshold_low = median_amount - (5 * mad)   # Lower threshold
                    
                    for entry in ledger_entries:
                        amount = float(entry.amount)
                        # Flag if amount is unusually large OR unusually small
                        if amount > threshold_high:
                            self.review_flags.append(
                                EntryReviewFlag(
                                    entry_id=entry.entry_id,
                                    status=ReviewStatus.UNUSUAL_AMOUNT,
                                    reason=f"Unusually large amount: {amount:,.2f} (median: {median_amount:,.2f})",
                                    severity="warning",
                                    source_entry_id=entry.source_entry_id,
                                )
                            )
                        elif amount < threshold_low:
                            self.review_flags.append(
                                EntryReviewFlag(
                                    entry_id=entry.entry_id,
                                    status=ReviewStatus.UNUSUAL_AMOUNT,
                                    reason=f"Unusually small amount: {amount:,.2f} (median: {median_amount:,.2f})",
                                    severity="warning",
                                    source_entry_id=entry.source_entry_id,
                                )
                            )

        # Check for multiple rules applied to same source entry
        if self.audit_trail:
            for record in self.audit_trail.records:
                if len(record.applied_rules) > 1:
                    self.review_flags.append(
                        EntryReviewFlag(
                            entry_id=record.entry_id,
                            status=ReviewStatus.MULTIPLE_RULES,
                            reason=f"Multiple rules ({len(record.applied_rules)}) matched this entry",
                            severity="info",
                            source_entry_id=record.entry_id,
                        )
                    )

    def _create_review_dashboard(
        self,
        wb: Workbook,
        ledger_entries: list[LedgerEntry],
        journal_entries: list[JournalEntry],
    ) -> None:
        """Create review dashboard sheet with summary and issues."""
        ws = wb.create_sheet("Review Dashboard", 0)

        # Title
        ws["A1"] = "Review Dashboard - Transformation Preview"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        row = 3

        # Summary Statistics
        ws[f"A{row}"] = "Summary Statistics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Count flags by severity
        flags_by_severity = defaultdict(int)
        flags_by_status = defaultdict(int)
        for flag in self.review_flags:
            flags_by_severity[flag.severity] += 1
            flags_by_status[flag.status.value] += 1

        # Get error detector statistics for alignment
        total_errors_detected = 0
        total_warnings_detected = 0
        if self.error_detector:
            total_errors_detected = len(self.error_detector.get_all_errors())
            total_warnings_detected = len(self.error_detector.get_all_warnings())

        stats = [
            ("Total Journal Entries", len(journal_entries)),
            ("Total Ledger Entries", len(ledger_entries)),
            ("Entries Needing Review", len(self.review_flags)),
            ("Errors Detected (All)", total_errors_detected),
            ("Warnings Detected (All)", total_warnings_detected),
            ("Critical Issues", flags_by_severity.get("critical", 0)),
            ("Error Flags", flags_by_severity.get("error", 0)),
            ("Warning Flags", flags_by_severity.get("warning", 0)),
            ("Info Items", flags_by_severity.get("info", 0)),
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            # Color code based on value
            if "Issues" in label or "Errors" in label or "Critical" in label:
                if value > 0:
                    ws[f"B{row}"].fill = PatternFill(
                        start_color=self.COLOR_ERROR,
                        end_color=self.COLOR_ERROR,
                        fill_type="solid",
                    )
            row += 1

        # Issues by Type
        row += 1
        ws[f"A{row}"] = "Issues by Type"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ["Issue Type", "Count", "Severity"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")
        row += 1

        # Issue type counts
        issue_type_labels = {
            "no_match": "No Matching Rule",
            "validation_error": "Validation Error",
            "validation_warning": "Validation Warning",
            "unusual_amount": "Unusual Amount",
            "missing_account": "Missing Account",
            "multiple_rules": "Multiple Rules",
        }

        for status_type, count in sorted(flags_by_status.items()):
            label = issue_type_labels.get(status_type, status_type.replace("_", " ").title())
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            
            # Get severity for this type
            severity = "info"
            for flag in self.review_flags:
                if flag.status.value == status_type:
                    severity = flag.severity
                    break
            
            ws.cell(row=row, column=3, value=severity.title())
            
            # Color code
            if severity == "critical":
                color = self.COLOR_CRITICAL
            elif severity == "error":
                color = self.COLOR_ERROR
            elif severity == "warning":
                color = self.COLOR_WARNING
            else:
                color = self.COLOR_INFO
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid",
                )
            
            row += 1

        # Quick Actions
        row += 1
        ws[f"A{row}"] = "Quick Actions"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        actions = [
            ("1. Review 'Preview Table' sheet", "See all ledger entries with visual flags"),
            ("2. Review 'Comparison View' sheet", "See Journal → Ledger side-by-side"),
            ("3. Review 'Flagged Entries' sheet", "See only entries needing review"),
        ]

        for action, description in actions:
            ws[f"A{row}"] = action
            ws[f"B{row}"] = description
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15

    def _create_preview_table(
        self, wb: Workbook, ledger_entries: list[LedgerEntry]
    ) -> None:
        """Create preview table sheet showing all ledger entries with flags."""
        ws = wb.create_sheet("Preview Table")

        # Title
        ws["A1"] = "Ledger Entries Preview (Final Output)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:M1")

        # Headers
        headers = [
            "Status",
            "Entry ID",
            "Ledger ID",
            "Account Code",
            "Account Path",
            "Description",
            "Amount",
            "Date",
            "Quarter",
            "Year",
            "Source Entry ID",
            "Rule Applied",
            "Review Reason",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Create lookup for flags
        flags_by_entry_id = {flag.entry_id: flag for flag in self.review_flags}

        # Write entries
        row = 4
        for entry in ledger_entries:
            flag = flags_by_entry_id.get(entry.entry_id)

            # Status column (icon)
            if flag:
                if flag.severity == "critical":
                    status_icon = self.ICON_CRITICAL
                    status_color = self.COLOR_CRITICAL
                elif flag.severity == "error":
                    status_icon = self.ICON_ERROR
                    status_color = self.COLOR_ERROR
                elif flag.severity == "warning":
                    status_icon = self.ICON_WARNING
                    status_color = self.COLOR_WARNING
                else:
                    status_icon = self.ICON_INFO
                    status_color = self.COLOR_INFO
            else:
                status_icon = self.ICON_OK
                status_color = self.COLOR_OK

            ws.cell(row=row, column=1, value=status_icon)
            ws.cell(row=row, column=2, value=entry.entry_id)
            ws.cell(row=row, column=3, value=entry.account_code)  # Ledger ID
            ws.cell(row=row, column=4, value=entry.account_code)
            ws.cell(row=row, column=5, value=entry.account_path)
            ws.cell(row=row, column=6, value=entry.description)
            ws.cell(row=row, column=7, value=float(entry.amount))
            ws.cell(row=row, column=8, value=entry.date.isoformat())
            ws.cell(row=row, column=9, value=entry.quarter)
            ws.cell(row=row, column=10, value=entry.year)
            ws.cell(row=row, column=11, value=entry.source_entry_id or "")
            ws.cell(row=row, column=12, value=entry.rule_applied or "")
            ws.cell(row=row, column=13, value=flag.reason if flag else "")

            # Apply row color based on status
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(
                    start_color=status_color,
                    end_color=status_color,
                    fill_type="solid",
                )
                cell.border = Border(
                    left=Side(style="thin", color=self.COLOR_BORDER),
                    right=Side(style="thin", color=self.COLOR_BORDER),
                    top=Side(style="thin", color=self.COLOR_BORDER),
                    bottom=Side(style="thin", color=self.COLOR_BORDER),
                )

            # Format amount
            amount_cell = ws.cell(row=row, column=7)
            amount_cell.number_format = "#,##0.00"

            # Format date
            date_cell = ws.cell(row=row, column=8)
            date_cell.number_format = "YYYY-MM-DD"

            row += 1

        # Auto-adjust column widths
        column_widths = {
            "A": 10,  # Status
            "B": 15,  # Entry ID
            "C": 12,  # Ledger ID
            "D": 15,  # Account Code
            "E": 40,  # Account Path
            "F": 30,  # Description
            "G": 15,  # Amount
            "H": 12,  # Date
            "I": 8,   # Quarter
            "J": 8,   # Year
            "K": 15,  # Source Entry ID
            "L": 15,  # Rule Applied
            "M": 50,  # Review Reason
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top rows
        ws.freeze_panes = "A4"

        # Enable filters
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"

    def _create_comparison_view(
        self,
        wb: Workbook,
        ledger_entries: list[LedgerEntry],
        journal_entries: list[JournalEntry],
    ) -> None:
        """Create side-by-side comparison view (Journal → Ledger)."""
        ws = wb.create_sheet("Comparison View")

        # Title
        ws["A1"] = "Journal Entry → Ledger Entry Comparison"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:O1")

        # Headers
        headers = [
            "Status",
            "Journal Entry ID",
            "Journal Description",
            "Journal Amount",
            "Journal Date",
            "Journal Type",
            "→",
            "Ledger Entry ID",
            "Account Code",
            "Account Path",
            "Ledger Amount",
            "Ledger Date",
            "Rule Applied",
            "Quarter",
            "Year",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Create lookup maps
        journal_by_id = {entry.entry_id: entry for entry in journal_entries}
        ledger_by_source = defaultdict(list)
        for entry in ledger_entries:
            if entry.source_entry_id:
                ledger_by_source[entry.source_entry_id].append(entry)

        flags_by_source = defaultdict(list)
        for flag in self.review_flags:
            if flag.source_entry_id:
                flags_by_source[flag.source_entry_id].append(flag)

        # Write comparison rows
        row = 4
        for journal_entry in journal_entries:
            ledger_entries_for_journal = ledger_by_source.get(
                journal_entry.entry_id, []
            )
            flags = flags_by_source.get(journal_entry.entry_id, [])

            # Determine status
            if flags:
                flag = flags[0]  # Use first flag
                if flag.severity == "critical":
                    status_icon = self.ICON_CRITICAL
                    status_color = self.COLOR_CRITICAL
                elif flag.severity == "error":
                    status_icon = self.ICON_ERROR
                    status_color = self.COLOR_ERROR
                elif flag.severity == "warning":
                    status_icon = self.ICON_WARNING
                    status_color = self.COLOR_WARNING
                else:
                    status_icon = self.ICON_INFO
                    status_color = self.COLOR_INFO
            else:
                status_icon = self.ICON_OK
                status_color = self.COLOR_OK

            # If no ledger entries, show journal entry only
            if not ledger_entries_for_journal:
                ws.cell(row=row, column=1, value=status_icon)
                ws.cell(row=row, column=2, value=journal_entry.entry_id)
                ws.cell(row=row, column=3, value=journal_entry.description)
                ws.cell(row=row, column=4, value=float(journal_entry.amount))
                ws.cell(row=row, column=5, value=journal_entry.date.isoformat())
                ws.cell(row=row, column=6, value=journal_entry.old_type or "")
                ws.cell(row=row, column=7, value="→")
                ws.cell(row=row, column=8, value="NO MATCH")
                ws.cell(row=row, column=8).font = Font(bold=True, color="FF0000")

                # Apply color
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(
                        start_color=status_color,
                        end_color=status_color,
                        fill_type="solid",
                    )

                row += 1
            else:
                # Show journal entry once, then all corresponding ledger entries
                first_row_for_journal = row
                for idx, ledger_entry in enumerate(ledger_entries_for_journal):
                    # Journal entry columns (only on first row)
                    if idx == 0:
                        ws.cell(row=row, column=1, value=status_icon)
                        ws.cell(row=row, column=2, value=journal_entry.entry_id)
                        ws.cell(row=row, column=3, value=journal_entry.description)
                        ws.cell(row=row, column=4, value=float(journal_entry.amount))
                        ws.cell(row=row, column=5, value=journal_entry.date.isoformat())
                        ws.cell(row=row, column=6, value=journal_entry.old_type or "")
                        ws.cell(row=row, column=7, value="→")
                    else:
                        # Leave journal columns empty (will merge later)
                        pass

                    # Ledger entry columns
                    ws.cell(row=row, column=8, value=ledger_entry.entry_id)
                    ws.cell(row=row, column=9, value=ledger_entry.account_code)
                    ws.cell(row=row, column=10, value=ledger_entry.account_path)
                    ws.cell(row=row, column=11, value=float(ledger_entry.amount))
                    ws.cell(row=row, column=12, value=ledger_entry.date.isoformat())
                    ws.cell(row=row, column=13, value=ledger_entry.rule_applied or "")
                    ws.cell(row=row, column=14, value=ledger_entry.quarter)
                    ws.cell(row=row, column=15, value=ledger_entry.year)

                    # Apply color
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is None:
                            continue
                        cell.fill = PatternFill(
                            start_color=status_color,
                            end_color=status_color,
                            fill_type="solid",
                        )

                    # Format amounts
                    amount_cell = ws.cell(row=row, column=4)
                    amount_cell.number_format = "#,##0.00"
                    amount_cell = ws.cell(row=row, column=11)
                    amount_cell.number_format = "#,##0.00"

                    # Format dates
                    date_cell = ws.cell(row=row, column=5)
                    date_cell.number_format = "YYYY-MM-DD"
                    date_cell = ws.cell(row=row, column=12)
                    date_cell.number_format = "YYYY-MM-DD"

                    row += 1

                # Merge journal entry columns for multi-ledger entries
                if len(ledger_entries_for_journal) > 1:
                    last_row_for_journal = row - 1
                    for merge_col in [1, 2, 3, 4, 5, 6, 7]:  # Status through arrow
                        if first_row_for_journal < last_row_for_journal:
                            ws.merge_cells(
                                start_row=first_row_for_journal,
                                start_column=merge_col,
                                end_row=last_row_for_journal,
                                end_column=merge_col,
                            )

        # Auto-adjust column widths
        column_widths = {
            "A": 10,  # Status
            "B": 15,  # Journal Entry ID
            "C": 30,  # Journal Description
            "D": 15,  # Journal Amount
            "E": 12,  # Journal Date
            "F": 15,  # Journal Type
            "G": 5,   # Arrow
            "H": 15,  # Ledger Entry ID
            "I": 15,  # Account Code
            "J": 40,  # Account Path
            "K": 15,  # Ledger Amount
            "L": 12,  # Ledger Date
            "M": 15,  # Rule Applied
            "N": 8,   # Quarter
            "O": 8,   # Year
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top rows
        ws.freeze_panes = "A4"

        # Enable filters
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"

    def _create_flagged_entries(
        self, wb: Workbook, ledger_entries: list[LedgerEntry], journal_entries: list[JournalEntry]
    ) -> None:
        """Create sheet showing only entries that need review."""
        ws = wb.create_sheet("Flagged Entries")

        # Title
        ws["A1"] = "Entries Requiring Review"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:M1")

        if not self.review_flags:
            ws["A3"] = "✓ No entries require review - all entries are OK!"
            ws["A3"].font = Font(bold=True, color="008000", size=12)
            return

        # Headers
        headers = [
            "Status",
            "Severity",
            "Entry ID",
            "Account Code",
            "Account Path",
            "Description",
            "Amount",
            "Date",
            "Source Entry ID",
            "Rule Applied",
            "Issue Type",
            "Reason",
            "Action Needed",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Create lookup for ledger entries and journal entries
        ledger_by_id = {entry.entry_id: entry for entry in ledger_entries}
        journal_by_id = {entry.entry_id: entry for entry in journal_entries}

        # Write flagged entries
        row = 4
        for flag in sorted(self.review_flags, key=lambda f: (
            {"critical": 0, "error": 1, "warning": 2, "info": 3}.get(f.severity, 4),
            f.entry_id
        )):
            entry = ledger_by_id.get(flag.entry_id)

            # Status icon
            if flag.severity == "critical":
                status_icon = self.ICON_CRITICAL
                status_color = self.COLOR_CRITICAL
            elif flag.severity == "error":
                status_icon = self.ICON_ERROR
                status_color = self.COLOR_ERROR
            elif flag.severity == "warning":
                status_icon = self.ICON_WARNING
                status_color = self.COLOR_WARNING
            else:
                status_icon = self.ICON_INFO
                status_color = self.COLOR_INFO

            ws.cell(row=row, column=1, value=status_icon)
            ws.cell(row=row, column=2, value=flag.severity.title())
            ws.cell(row=row, column=3, value=flag.entry_id)
            
            if entry:
                ws.cell(row=row, column=4, value=entry.account_code or "")
                ws.cell(row=row, column=5, value=entry.account_path or "")
                ws.cell(row=row, column=6, value=entry.description or "")
                ws.cell(row=row, column=7, value=float(entry.amount) if entry.amount else 0.0)
                ws.cell(row=row, column=8, value=entry.date.isoformat() if entry.date else "")
                ws.cell(row=row, column=9, value=entry.source_entry_id or "")
                ws.cell(row=row, column=10, value=entry.rule_applied or "")
            else:
                # Entry not found in ledger (might be a journal entry with no match)
                # Try to find the journal entry
                journal_entry = journal_by_id.get(flag.entry_id)
                if not journal_entry and flag.source_entry_id:
                    # Try source_entry_id
                    journal_entry = journal_by_id.get(flag.source_entry_id)
                
                if journal_entry:
                    ws.cell(row=row, column=4, value="N/A (Journal Entry)")
                    ws.cell(row=row, column=5, value="N/A (Journal Entry)")
                    ws.cell(row=row, column=6, value=journal_entry.description or "Journal entry - no ledger match")
                    ws.cell(row=row, column=7, value=float(journal_entry.amount) if journal_entry.amount else 0.0)
                    ws.cell(row=row, column=8, value=journal_entry.date.isoformat() if journal_entry.date else "")
                    ws.cell(row=row, column=9, value=flag.source_entry_id or flag.entry_id or "")
                    ws.cell(row=row, column=10, value="N/A (No transformation)")
                else:
                    ws.cell(row=row, column=4, value="N/A")
                    ws.cell(row=row, column=5, value="N/A")
                    ws.cell(row=row, column=6, value="Entry not found in ledger or journal")
                    ws.cell(row=row, column=7, value="N/A")
                    ws.cell(row=row, column=8, value="N/A")
                    ws.cell(row=row, column=9, value=flag.source_entry_id or flag.entry_id or "")
                    ws.cell(row=row, column=10, value="N/A")

            ws.cell(row=row, column=11, value=flag.status.value.replace("_", " ").title())
            ws.cell(row=row, column=12, value=flag.reason)
            
            # Action needed
            action_map = {
                ReviewStatus.NO_MATCH: "Review journal entry and add/update mapping rule",
                ReviewStatus.VALIDATION_ERROR: "Fix data error in journal entry",
                ReviewStatus.VALIDATION_WARNING: "Review warning - may need correction",
                ReviewStatus.UNUSUAL_AMOUNT: "Verify amount is correct",
                ReviewStatus.MISSING_ACCOUNT: "Verify account code or update hierarchy",
                ReviewStatus.MULTIPLE_RULES: "Review rule priority - may need adjustment",
            }
            ws.cell(row=row, column=13, value=action_map.get(flag.status, "Review entry"))

            # Apply row color
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(
                    start_color=status_color,
                    end_color=status_color,
                    fill_type="solid",
                )
                cell.border = Border(
                    left=Side(style="thin", color=self.COLOR_BORDER),
                    right=Side(style="thin", color=self.COLOR_BORDER),
                    top=Side(style="thin", color=self.COLOR_BORDER),
                    bottom=Side(style="thin", color=self.COLOR_BORDER),
                )

            # Format amount
            if entry:
                amount_cell = ws.cell(row=row, column=7)
                amount_cell.number_format = "#,##0.00"
                date_cell = ws.cell(row=row, column=8)
                date_cell.number_format = "YYYY-MM-DD"

            row += 1

        # Auto-adjust column widths
        column_widths = {
            "A": 10,  # Status
            "B": 12,  # Severity
            "C": 15,  # Entry ID
            "D": 15,  # Account Code
            "E": 40,  # Account Path
            "F": 30,  # Description
            "G": 15,  # Amount
            "H": 12,  # Date
            "I": 15,  # Source Entry ID
            "J": 15,  # Rule Applied
            "K": 20,  # Issue Type
            "L": 50,  # Reason
            "M": 40,  # Action Needed
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top rows
        ws.freeze_panes = "A4"

        # Enable filters
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"
