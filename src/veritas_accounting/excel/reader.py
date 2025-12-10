"""Excel file reading module for veritas-accounting."""

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from veritas_accounting.utils.exceptions import ExcelIOError


class ExcelReader:
    """Reads Excel files using openpyxl with UTF-8 encoding support."""

    def __init__(self):
        """Initialize the Excel reader."""
        pass

    def read_file(
        self,
        file_path: str | Path,
        sheet_name: Optional[str] = None,
        header: int = 0,
        engine: str = "openpyxl",
    ) -> pd.DataFrame:
        """
        Read an Excel file and return data as a pandas DataFrame.

        Args:
            file_path: Path to the Excel file (.xlsx format)
            sheet_name: Name of the worksheet to read. If None, reads the first sheet.
            header: Row number to use as column names (0-indexed). Default is 0.
            engine: Engine to use for reading. Default is 'openpyxl'.

        Returns:
            pandas DataFrame containing the worksheet data

        Raises:
            ExcelIOError: If file is missing, corrupted, or has invalid format
        """
        file_path = Path(file_path)

        # Check if file exists
        if not file_path.exists():
            raise ExcelIOError(
                f"Excel file not found: {file_path}. "
                "Please check that the file path is correct and the file exists."
            )

        # Check if file is a directory
        if file_path.is_dir():
            raise ExcelIOError(
                f"Path is a directory, not a file: {file_path}. "
                "Please provide a valid Excel file path."
            )

        # Validate file extension
        if file_path.suffix.lower() != ".xlsx":
            raise ExcelIOError(
                f"Invalid file format: {file_path}. "
                "Only .xlsx files are supported. Please convert your file to .xlsx format."
            )

        try:
            # Read Excel file using pandas with openpyxl engine
            # pandas handles UTF-8 encoding automatically for Excel files
            result = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header,
                engine=engine,
            )

            # When sheet_name is None, pandas returns a dict of DataFrames
            # We want to return the first sheet as a DataFrame
            if isinstance(result, dict):
                # Get the first sheet
                first_sheet_name = list(result.keys())[0]
                df = result[first_sheet_name]
            else:
                df = result

            # Handle empty worksheets (return empty DataFrame)
            return df

        except FileNotFoundError:
            # This should not happen since we check above, but handle it anyway
            raise ExcelIOError(
                f"Excel file not found: {file_path}. "
                "Please check that the file path is correct."
            )
        except InvalidFileException as e:
            raise ExcelIOError(
                f"Invalid Excel file format: {file_path}. "
                f"The file may be corrupted or not a valid .xlsx file. "
                f"Original error: {str(e)}"
            )
        except Exception as e:
            # Catch any other openpyxl or pandas exceptions
            error_msg = str(e)
            if "No such file" in error_msg or "cannot find" in error_msg.lower():
                raise ExcelIOError(
                    f"Excel file not found: {file_path}. "
                    "Please check that the file path is correct."
                )
            elif "corrupt" in error_msg.lower() or "invalid" in error_msg.lower():
                raise ExcelIOError(
                    f"Corrupted Excel file: {file_path}. "
                    f"The file cannot be read. Original error: {error_msg}"
                )
            else:
                raise ExcelIOError(
                    f"Error reading Excel file: {file_path}. "
                    f"Original error: {error_msg}"
                )

    def read_sheet(
        self,
        file_path: str | Path,
        sheet_name: str,
        header: int = 0,
        engine: str = "openpyxl",
    ) -> pd.DataFrame:
        """
        Read a specific worksheet from an Excel file.

        Args:
            file_path: Path to the Excel file (.xlsx format)
            sheet_name: Name of the worksheet to read
            header: Row number to use as column names (0-indexed). Default is 0.
            engine: Engine to use for reading. Default is 'openpyxl'.

        Returns:
            pandas DataFrame containing the worksheet data

        Raises:
            ExcelIOError: If file is missing, corrupted, sheet doesn't exist, or has invalid format
        """
        file_path = Path(file_path)

        # Check if file exists
        if not file_path.exists():
            raise ExcelIOError(
                f"Excel file not found: {file_path}. "
                "Please check that the file path is correct and the file exists."
            )

        # Validate file extension
        if file_path.suffix.lower() != ".xlsx":
            raise ExcelIOError(
                f"Invalid file format: {file_path}. "
                "Only .xlsx files are supported."
            )

        try:
            # First, verify the sheet exists
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                raise ExcelIOError(
                    f"Worksheet '{sheet_name}' not found in {file_path}. "
                    f"Available worksheets: {available_sheets}"
                )
            workbook.close()

            # Read the specific sheet using pandas
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header,
                engine=engine,
            )

            # Handle empty worksheets (return empty DataFrame)
            return df

        except FileNotFoundError:
            raise ExcelIOError(
                f"Excel file not found: {file_path}. "
                "Please check that the file path is correct."
            )
        except InvalidFileException as e:
            raise ExcelIOError(
                f"Invalid Excel file format: {file_path}. "
                f"The file may be corrupted or not a valid .xlsx file. "
                f"Original error: {str(e)}"
            )
        except ExcelIOError:
            # Re-raise our custom exceptions
            raise
        except Exception as e:
            error_msg = str(e)
            if "No such file" in error_msg or "cannot find" in error_msg.lower():
                raise ExcelIOError(
                    f"Excel file not found: {file_path}. "
                    "Please check that the file path is correct."
                )
            elif "corrupt" in error_msg.lower() or "invalid" in error_msg.lower():
                raise ExcelIOError(
                    f"Corrupted Excel file: {file_path}. "
                    f"The file cannot be read. Original error: {error_msg}"
                )
            else:
                raise ExcelIOError(
                    f"Error reading worksheet '{sheet_name}' from {file_path}. "
                    f"Original error: {error_msg}"
                )

    def list_sheets(self, file_path: str | Path) -> list[str]:
        """
        List all worksheet names in an Excel file.

        Args:
            file_path: Path to the Excel file (.xlsx format)

        Returns:
            List of worksheet names

        Raises:
            ExcelIOError: If file is missing, corrupted, or has invalid format
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise ExcelIOError(
                f"Excel file not found: {file_path}. "
                "Please check that the file path is correct and the file exists."
            )

        if file_path.suffix.lower() != ".xlsx":
            raise ExcelIOError(
                f"Invalid file format: {file_path}. "
                "Only .xlsx files are supported."
            )

        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except InvalidFileException as e:
            raise ExcelIOError(
                f"Invalid Excel file format: {file_path}. "
                f"The file may be corrupted or not a valid .xlsx file. "
                f"Original error: {str(e)}"
            )
        except Exception as e:
            raise ExcelIOError(
                f"Error reading Excel file: {file_path}. "
                f"Original error: {str(e)}"
            )
