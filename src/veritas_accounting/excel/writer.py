"""Excel file writing with formatting support for veritas-accounting."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from veritas_accounting.utils.exceptions import ExcelIOError


class ExcelWriter:
    """Excel file writer with formatting support."""

    def __init__(self) -> None:
        """Initialize ExcelWriter."""
        pass

    def write_file(
        self,
        data: pd.DataFrame | list[dict[str, Any]],
        path: Path | str,
        sheet_name: str = "Sheet1",
        overwrite: bool = True,
    ) -> None:
        """
        Write data to an Excel file.

        Args:
            data: DataFrame or list of dictionaries to write
            path: Output file path
            sheet_name: Name of the worksheet (default: "Sheet1")
            overwrite: If True, overwrite existing file; if False, append to existing file

        Raises:
            ExcelIOError: If writing fails (permissions, invalid data, etc.)
        """
        path = Path(path)

        # Create directory structure if it doesn't exist
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            raise ExcelIOError(
                f"Failed to create output directory: {path.parent}. Error: {e}"
            ) from e

        # Convert list of dicts to DataFrame if needed
        if isinstance(data, list):
            if not data:
                df = pd.DataFrame()
            else:
                df = pd.DataFrame(data)
        else:
            df = data

        # Validate data format
        if not isinstance(df, pd.DataFrame):
            raise ExcelIOError(
                f"Invalid data format. Expected DataFrame or list of dicts, got {type(data)}"
            )

        try:
            # Write to Excel using pandas with openpyxl engine
            if overwrite or not path.exists():
                # Create new file
                with pd.ExcelWriter(
                    path, engine="openpyxl", mode="w"
                ) as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Append to existing file
                with pd.ExcelWriter(
                    path, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply formatting using openpyxl
            self._apply_formatting(path, sheet_name)

        except PermissionError as e:
            raise ExcelIOError(
                f"Permission denied: Cannot write to {path}. "
                "File may be open in another application."
            ) from e
        except Exception as e:
            raise ExcelIOError(
                f"Failed to write Excel file: {path}. Error: {e}"
            ) from e

    def write_sheet(
        self,
        data: pd.DataFrame | list[dict[str, Any]],
        path: Path | str,
        sheet_name: str,
        overwrite: bool = True,
    ) -> None:
        """
        Write data to a specific worksheet in an Excel file.

        Args:
            data: DataFrame or list of dictionaries to write
            path: Output file path
            sheet_name: Name of the worksheet
            overwrite: If True, replace existing sheet; if False, append data to existing sheet

        Raises:
            ExcelIOError: If writing fails
        """
        self.write_file(data, path, sheet_name, overwrite)

    def _apply_formatting(
        self, path: Path, sheet_name: str
    ) -> None:
        """
        Apply formatting to the Excel file (column widths, header styling).

        Args:
            path: Path to the Excel file
            sheet_name: Name of the worksheet to format
        """
        try:
            workbook = load_workbook(path)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return

            worksheet = workbook[sheet_name]

            # Format header row (first row)
            if worksheet.max_row > 0:
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font

            # Auto-adjust column widths
            self._auto_adjust_column_widths(worksheet)

            workbook.save(path)
            workbook.close()

        except Exception as e:
            # Don't fail if formatting fails, just log it
            # The file was already written successfully
            raise ExcelIOError(
                f"Failed to apply formatting to {path}. Error: {e}"
            ) from e

    def _auto_adjust_column_widths(self, worksheet: Any) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: openpyxl worksheet object
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    # Get cell value length
                    if cell.value:
                        cell_value = str(cell.value)
                        # Handle Chinese characters (they take more width)
                        # Rough estimate: Chinese chars are ~2x width
                        chinese_chars = sum(
                            1 for c in cell_value if ord(c) > 127
                        )
                        length = len(cell_value) + chinese_chars
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass

            # Set column width with some padding
            # Minimum width 10, maximum 50, add 2 for padding
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def write_multiple_sheets(
        self,
        sheets: dict[str, pd.DataFrame | list[dict[str, Any]]],
        path: Path | str,
        overwrite: bool = True,
    ) -> None:
        """
        Write multiple sheets to a single Excel file.

        Args:
            sheets: Dictionary mapping sheet names to DataFrames or lists of dicts
            path: Output file path
            overwrite: If True, overwrite existing file

        Raises:
            ExcelIOError: If writing fails
        """
        path = Path(path)

        # Create directory structure if it doesn't exist
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            raise ExcelIOError(
                f"Failed to create output directory: {path.parent}. Error: {e}"
            ) from e

        try:
            # Write all sheets using pandas
            with pd.ExcelWriter(
                path, engine="openpyxl", mode="w" if overwrite else "a"
            ) as writer:
                for sheet_name, data in sheets.items():
                    # Convert list of dicts to DataFrame if needed
                    if isinstance(data, list):
                        if not data:
                            df = pd.DataFrame()
                        else:
                            df = pd.DataFrame(data)
                    else:
                        df = data

                    df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                    )

            # Apply formatting to all sheets
            workbook = load_workbook(path)
            for sheet_name in sheets.keys():
                if sheet_name in workbook.sheetnames:
                    self._apply_formatting_to_workbook(workbook, sheet_name)
            workbook.save(path)
            workbook.close()

        except PermissionError as e:
            raise ExcelIOError(
                f"Permission denied: Cannot write to {path}. "
                "File may be open in another application."
            ) from e
        except Exception as e:
            raise ExcelIOError(
                f"Failed to write Excel file with multiple sheets: {path}. Error: {e}"
            ) from e

    def _apply_formatting_to_workbook(
        self, workbook: Workbook, sheet_name: str
    ) -> None:
        """
        Apply formatting to a worksheet in an open workbook.

        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name of the worksheet to format
        """
        if sheet_name not in workbook.sheetnames:
            return

        worksheet = workbook[sheet_name]

        # Format header row
        if worksheet.max_row > 0:
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        # Auto-adjust column widths
        self._auto_adjust_column_widths(worksheet)








